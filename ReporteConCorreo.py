import streamlit as st
import polars as pl
import pandas as pd
import io
import html
import json
import urllib.parse
import streamlit.components.v1 as components
import datetime as _dt
from datetime import date

st.set_page_config(
    page_title="Seguimiento Regalías",
    layout="wide",
    initial_sidebar_state="expanded",
)

C = {
    "azul_oscuro":  "#003d6c",
    "azul_medio":   "#1754ab",
    "verde_oscuro": "#005931",
    "verde_medio":  "#17743d",
    "cian":         "#47b1d5",
    "naranja":      "#d88c16",
    "naranja_osc":  "#cf7000",
    "cafe":         "#9b5b1e",
    "salmon":       "#e68878",
    "bg":           "#e8edf5",
    "white":        "#ffffff",
    "text":         "#1a2332",
    "muted":        "#6b7280",
    "border":       "#e2e8f0",
}

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {{
    font-family: 'Montserrat', sans-serif;
    color: {C['text']};
}}

/* ── Fondo del área principal con gradiente institucional ── */
.stApp {{
    background:
        radial-gradient(ellipse at 0% 0%, rgba(0,61,108,0.10) 0%, transparent 55%),
        radial-gradient(ellipse at 100% 100%, rgba(0,89,49,0.08) 0%, transparent 55%),
        linear-gradient(160deg, #dde5f0 0%, #e8edf5 40%, #eaf0f0 100%);
    min-height: 100vh;
}}

/* Controlar padding del contenedor principal de Streamlit */
.block-container {{
    padding-top: 3.5rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 1400px !important;
}}

/* ── Sidebar oscuro ── */
section[data-testid="stSidebar"] > div {{
    background: #001f3f;
    padding-top: 1.5rem;
}}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span:not([data-baseweb="tag"] span),
section[data-testid="stSidebar"] div {{
    color: rgba(255,255,255,0.95) !important;
}}
/* File uploader — caja blanca con texto oscuro */
section[data-testid="stSidebar"] [data-testid="stFileUploader"] section {{
    background: rgba(255,255,255,0.08) !important;
    border: 1.5px dashed rgba(255,255,255,0.25) !important;
    border-radius: 8px !important;
}}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] section *{{
    color: rgba(255,255,255,0.85) !important;
}}
section[data-testid="stSidebar"] [data-testid="stFileUploader"] button {{
    background: rgba(255,255,255,0.12) !important;
    color: white !important;
    border: 1px solid rgba(255,255,255,0.3) !important;
    border-radius: 6px !important;
}}
/* Tags del multiselect */
span[data-baseweb="tag"] {{
    background: {C['azul_medio']} !important;
    color: white !important;
    max-width: 160px !important;
    border-radius: 4px !important;
}}
span[data-baseweb="tag"] span {{
    color: white !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    overflow: hidden !important;
    white-space: nowrap !important;
    text-overflow: ellipsis !important;
    max-width: 130px !important;
    display: block !important;
}}
/* Contenedor multiselect en sidebar */
section[data-testid="stSidebar"] [data-baseweb="select"] > div:first-child {{
    flex-wrap: wrap !important;
    gap: 4px !important;
    padding: 6px 8px !important;
    background: rgba(255,255,255,0.07) !important;
    border-color: rgba(255,255,255,0.18) !important;
    border-radius: 6px !important;
}}
/* Selectbox valor visible */
section[data-testid="stSidebar"] [data-baseweb="select"] [class*="singleValue"],
section[data-testid="stSidebar"] [data-baseweb="select"] [class*="placeholder"] {{
    color: white !important;
    font-weight: 500 !important;
}}
/* Labels de filtros */
section[data-testid="stSidebar"] .stMultiSelect label,
section[data-testid="stSidebar"] .stSelectbox label {{
    color: rgba(255,255,255,0.6) !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
    margin-bottom: 0.2rem !important;
}}
/* Sección título sidebar */
.sidebar-section {{
    font-family: 'Montserrat', sans-serif;
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    color: {C['cian']} !important;
    margin: 1.5rem 0 0.7rem 0;
    padding-bottom: 0.4rem;
    border-bottom: 1px solid rgba(71,177,213,0.3);
}}

/* ── Header ── */
.page-header {{
    background: linear-gradient(120deg, {C['azul_oscuro']} 0%, {C['verde_oscuro']} 100%);
    border-radius: 12px;
    margin: 0 0 0.8rem 0;
    padding: 1.8rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 4px 24px rgba(0,61,108,0.22);
    position: relative;
    z-index: 1;
}}
.page-header h1 {{
    font-family: 'Montserrat', sans-serif;
    font-size: 1.6rem;
    font-weight: 800;
    color: white;
    margin: 0 0 0.15rem 0;
}}
.page-header p {{
    color: rgba(255,255,255,0.6);
    margin: 0;
    font-size: 0.82rem;
    font-weight: 400;
}}

/* ── Filtros horizontales ── */
.filter-bar {{
    background: {C['white']};
    border-radius: 10px;
    padding: 0.85rem 1.2rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    display: flex;
    align-items: center;
    gap: 1rem;
    flex-wrap: wrap;
}}
.filter-label {{
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: {C['muted']};
    white-space: nowrap;
}}

/* ── KPI principal ── */
.kpi-main {{
    background: {C['azul_oscuro']};
    border-radius: 12px;
    padding: 1.4rem 1.6rem;
    color: white;
    height: 100%;
    box-shadow: 0 4px 20px rgba(0,61,108,0.22), 0 1px 4px rgba(0,0,0,0.1);
}}
.kpi-main .label {{
    font-size: 0.68rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: rgba(255,255,255,0.6);
    margin-bottom: 0.4rem;
}}
.kpi-main .value {{
    font-family: 'Montserrat', sans-serif;
    font-size: 2.8rem;
    font-weight: 800;
    line-height: 1;
    color: white;
}}
.kpi-main .sub {{
    font-size: 0.78rem;
    color: rgba(255,255,255,0.5);
    margin-top: 0.2rem;
}}

/* ── KPI secundario ── */
.kpi-sec {{
    background: #ffffff;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    border-left: 4px solid;
    box-shadow: 0 3px 16px rgba(0,40,90,0.13), 0 1px 3px rgba(0,0,0,0.07);
    height: 100%;
}}
.kpi-sec .label {{
    font-size: 0.63rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: {C['muted']};
    margin-bottom: 0.3rem;
}}
.kpi-sec .value {{
    font-family: 'Montserrat', sans-serif;
    font-size: 1.6rem;
    font-weight: 800;
    line-height: 1;
}}
.kpi-sec .sub {{
    font-size: 0.7rem;
    color: {C['muted']};
    margin-top: 0.15rem;
}}
/* Columna con dos kpi-sec apiladas */
.kpi-stack {{
    display: flex;
    flex-direction: column;
    height: 100%;
    gap: 0;
}}
.kpi-stack .kpi-sec {{
    flex: 1;
    height: auto;
    padding: 0.7rem 1.2rem;
}}
/* Tarjeta de conteo por estado */
.kpi-estados {{
    background: #ffffff;
    border-radius: 10px;
    padding: 0.85rem 1.1rem;
    box-shadow: 0 3px 16px rgba(0,40,90,0.13), 0 1px 3px rgba(0,0,0,0.07);
    height: 100%;
}}
.kpi-estados-title {{
    font-size: 0.63rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: {C['muted']};
    margin-bottom: 0.55rem;
}}
.kpi-estados-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 0 1.2rem;
}}
.estado-kpi-row {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 0.28rem 0;
    border-bottom: 1px solid {C['border']};
    font-size: 0.78rem;
}}
.estado-kpi-row:last-child {{ border-bottom: none; }}
.estado-kpi-label {{
    display: flex;
    align-items: center;
    gap: 6px;
    color: {C['muted']};
    font-weight: 500;
    font-size: 0.72rem;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 80%;
    text-transform: uppercase;
    letter-spacing: 0.3px;
}}
.estado-kpi-n {{
    font-family: 'DM Mono', monospace;
    font-weight: 800;
    font-size: 1.05rem;
    color: {C['azul_oscuro']};
    min-width: 2rem;
    text-align: right;
    line-height: 1;
}}

/* ── Section heading ── */
.section-heading {{
    font-family: 'Montserrat', sans-serif;
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    color: {C['azul_oscuro']};
    border-bottom: 2px solid {C['cian']};
    padding-bottom: 0.45rem;
    margin: 0 0 1rem 0;
}}

/* ── Tabs ── */
div[data-testid="stTabs"] [role="tablist"] {{
    background: rgba(255,255,255,0.70);
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-radius: 10px 10px 0 0;
    padding: 0 1rem;
    border-bottom: 2px solid {C['border']};
    gap: 0;
}}
div[data-testid="stTabs"] [role="tab"] {{
    font-family: 'Montserrat', sans-serif !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    color: {C['muted']} !important;
    padding: 0.75rem 1.2rem !important;
    border-bottom: 2px solid transparent !important;
    margin-bottom: -2px !important;
}}
div[data-testid="stTabs"] [role="tab"][aria-selected="true"] {{
    color: {C['azul_oscuro']} !important;
    border-bottom-color: {C['cian']} !important;
}}
div[data-testid="stTabs"] [data-testid="stTabsContent"] {{
    background: #ffffff;
    border-radius: 0 0 12px 12px;
    padding: 1.4rem;
    box-shadow: 0 4px 24px rgba(0,40,90,0.10), 0 1px 4px rgba(0,0,0,0.06);
}}

/* ── Summary table ── */
.summary-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.83rem;
    background: #ffffff;
    border-radius: 8px;
    overflow: hidden;
}}
.summary-table thead tr {{
    background: {C['azul_oscuro']};
    color: white;
}}
.summary-table th {{
    padding: 0.7rem 0.9rem;
    font-family: 'Montserrat', sans-serif;
    font-size: 0.64rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    text-align: left;
    white-space: nowrap;
    line-height: 1.4;
}}
.summary-table td {{
    padding: 0.65rem 0.9rem;
    border-bottom: 1px solid {C['border']};
    vertical-align: middle;
    background: #ffffff;
}}
.summary-table tbody tr:nth-child(even) td {{ background: #f7fafd; }}
.summary-table tbody tr:last-child td {{ border-bottom: none; }}
.summary-table tbody tr:hover td {{ background: #e8f3ff !important; transition: background 0.15s; }}
/* Columna Total destacada */
.summary-table .col-total {{
    background: {C['azul_oscuro']} !important;
    font-family: 'DM Mono', monospace;
    font-weight: 700;
    font-size: 0.83rem;
    color: #ffffff !important;
    border-left: 2px solid rgba(255,255,255,0.2);
    text-align: center;
}}
.summary-table thead .col-total {{
    background: {C['azul_medio']} !important;
    font-size: 0.64rem;
    font-family: 'Montserrat', sans-serif;
    font-weight: 700;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}
.summary-table tbody tr:nth-child(even) .col-total {{ background: {C['azul_oscuro']} !important; }}
.summary-table tbody tr:hover .col-total {{ background: {C['azul_medio']} !important; }}
.entidad-name {{
    font-weight: 600;
    font-size: 0.83rem;
    color: {C['azul_oscuro']};
}}
.dias-val {{
    font-family: 'DM Mono', monospace;
    font-size: 0.8rem;
    font-weight: 500;
}}
.null-cell {{ color: {C['border']}; }}

/* ── Badges ── */
.badge {{
    display: inline-flex;
    align-items: center;
    gap: 5px;
    padding: 3px 9px;
    border-radius: 20px;
    font-size: 0.68rem;
    font-weight: 700;
    margin-left: 5px;
    vertical-align: middle;
    cursor: help;
    position: relative;
}}
.badge-green  {{ background: #d1fae5; color: #065f46; }}
.badge-yellow {{ background: #fef3c7; color: #92400e; }}
.badge-orange {{ background: #ffedd5; color: #9a3412; }}
.badge-red    {{ background: #fee2e2; color: #991b1b; }}
.badge-black  {{ background: #1e293b; color: #f1f5f9; }}
/* Punto de color semáforo */
.badge-dot {{
    width: 8px; height: 8px;
    border-radius: 50%;
    flex-shrink: 0;
    display: inline-block;
}}
.badge-green  .badge-dot {{ background: #059669; }}
.badge-yellow .badge-dot {{ background: #d97706; }}
.badge-orange .badge-dot {{ background: #ea580c; }}
.badge-red    .badge-dot {{ background: #dc2626; }}
.badge-black  .badge-dot {{ background: #94a3b8; }}
/* Tooltip del semáforo */
.badge-tooltip {{
    display: none;
    position: fixed;
    background: {C['text']};
    color: white;
    font-family: 'Montserrat', sans-serif;
    font-size: 0.72rem;
    font-weight: 400;
    line-height: 1.5;
    padding: 0.55rem 0.8rem;
    border-radius: 8px;
    width: 240px;
    text-align: left;
    text-transform: none;
    letter-spacing: 0;
    z-index: 99999;
    box-shadow: 0 4px 16px rgba(0,0,0,0.25);
    pointer-events: none;
    white-space: normal;
}}
.badge-tooltip::after {{
    content: '';
    position: absolute;
    left: 50%;
    transform: translateX(-50%);
    border: 5px solid transparent;
}}
.badge-tooltip.tip-arriba::after {{
    top: 100%;
    border-top-color: {C['text']};
}}
.badge-tooltip.tip-abajo::after {{
    bottom: 100%;
    border-bottom-color: {C['text']};
}}

/* ── Evaluación — calificación card ── */
.eval-card {{
    background: #ffffff;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    box-shadow: 0 2px 12px rgba(0,40,90,0.09);
    margin-bottom: 0.6rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}}
.eval-bar-wrap {{
    flex: 1;
}}
.eval-label {{
    font-size: 0.72rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: {C['muted']};
    margin-bottom: 0.3rem;
}}
.eval-bar-bg {{
    background: {C['border']};
    border-radius: 6px;
    height: 10px;
    overflow: hidden;
    margin-bottom: 0.2rem;
}}
.eval-bar-fill {{
    height: 100%;
    border-radius: 6px;
    transition: width 0.4s ease;
}}
.eval-score {{
    font-family: 'DM Mono', monospace;
    font-size: 1.1rem;
    font-weight: 700;
    min-width: 3rem;
    text-align: right;
    color: {C['azul_oscuro']};
}}

/* ── Detail table ── */
.detail-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.81rem;
    margin-top: 0.4rem;
}}
.detail-table th {{
    background: #f1f5f9;
    color: {C['azul_oscuro']};
    font-family: 'Montserrat', sans-serif;
    font-size: 0.63rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    padding: 0.5rem 0.85rem;
    text-align: left;
    border-bottom: 2px solid {C['border']};
}}
.detail-table td {{
    padding: 0.55rem 0.85rem;
    border-bottom: 1px solid {C['border']};
    vertical-align: middle;
}}
.detail-table tbody tr:last-child td {{ border-bottom: none; }}
/* Relleno de fila según semáforo — tonos suaves para mantener legibilidad */
.detail-table tbody tr.row-green  td {{ background: #f0fdf4; }}
.detail-table tbody tr.row-yellow td {{ background: #fffbeb; }}
.detail-table tbody tr.row-orange td {{ background: #fff7ed; }}
.detail-table tbody tr.row-black  td {{ background: #f1f5f9; }}
/* Hover — tono ligeramente más intenso manteniendo el color del semáforo */
.detail-table tbody tr.row-green:hover  td {{ background: #dcfce7 !important; }}
.detail-table tbody tr.row-yellow:hover td {{ background: #fef3c7 !important; }}
.detail-table tbody tr.row-orange:hover td {{ background: #ffedd5 !important; }}
.detail-table tbody tr.row-black:hover  td {{ background: #e2e8f0 !important; }}
.detail-table tbody tr:hover td {{ background: #f0f6ff; }}
.bpin-tag {{
    font-family: 'DM Mono', monospace;
    font-size: 0.7rem;
    background: #f1f5f9;
    color: {C['muted']};
    padding: 2px 6px;
    border-radius: 4px;
}}
.estado-tag {{
    font-size: 0.7rem;
    background: #eff6ff;
    color: {C['azul_medio']};
    padding: 2px 7px;
    border-radius: 4px;
    font-weight: 500;
    white-space: nowrap;
}}

/* ── Expander ── */
div[data-testid="stExpander"] > details {{
    border-radius: 10px !important;
    overflow: hidden !important;
    box-shadow: 0 2px 10px rgba(0,40,90,0.08) !important;
    margin-bottom: 0.5rem !important;
    border: none !important;
}}
div[data-testid="stExpander"] > details > summary {{
    background: #ffffff !important;
    border: none !important;
    border-left: 4px solid {C['azul_oscuro']} !important;
    border-radius: 10px !important;
    font-family: 'Montserrat', sans-serif !important;
    font-weight: 700 !important;
    font-size: 0.84rem !important;
    color: {C['azul_oscuro']} !important;
    padding: 0.75rem 1.1rem !important;
    transition: background 0.15s, border-color 0.15s !important;
}}
div[data-testid="stExpander"] > details > summary:hover {{
    background: #f0f7ff !important;
    border-left-color: {C['cian']} !important;
}}
div[data-testid="stExpander"] > details[open] > summary {{
    border-radius: 10px 10px 0 0 !important;
    border-left-color: {C['cian']} !important;
    background: #f8fbff !important;
}}
div[data-testid="stExpander"] > details > div {{
    border: none !important;
    border-left: 4px solid {C['cian']} !important;
    border-radius: 0 0 10px 10px !important;
    background: #ffffff !important;
    padding: 0.2rem 1rem 0.8rem 1rem !important;
}}

/* ── Tooltip ── */
.th-wrap {{
    display: inline-flex;
    align-items: center;
    gap: 4px;
    cursor: default;
    position: relative;
}}
.th-icon {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 14px;
    height: 14px;
    border-radius: 50%;
    background: rgba(255,255,255,0.2);
    color: white;
    font-size: 0.6rem;
    font-weight: 700;
    cursor: help;
    line-height: 1;
    flex-shrink: 0;
}}
.th-tooltip {{
    visibility: hidden;
    opacity: 0;
    position: absolute;
    top: calc(100% + 8px);
    left: 50%;
    transform: translateX(-50%);
    background: {C['text']};
    color: white;
    font-family: 'Montserrat', sans-serif;
    font-size: 0.73rem;
    font-weight: 400;
    line-height: 1.55;
    padding: 0.6rem 0.85rem;
    border-radius: 8px;
    white-space: normal;
    width: 230px;
    text-align: left;
    text-transform: none;
    letter-spacing: 0;
    z-index: 9999;
    box-shadow: 0 4px 18px rgba(0,0,0,0.28);
    pointer-events: none;
    transition: opacity 0.15s ease;
}}
.th-tooltip strong {{
    display: block;
    margin-bottom: 4px;
    font-size: 0.75rem;
    color: {C['cian']};
}}
.th-tooltip::before {{
    content: '';
    position: absolute;
    top: -6px;
    left: 50%;
    transform: translateX(-50%);
    border: 6px solid transparent;
    border-top: none;
    border-bottom-color: {C['text']};
}}
.th-wrap:hover .th-tooltip {{ visibility: visible; opacity: 1; }}

/* ── Download button ── */
.stDownloadButton > button {{
    background: {C['verde_oscuro']} !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-family: 'Montserrat', sans-serif !important;
    font-size: 0.83rem !important;
    padding: 0.5rem 1.4rem !important;
}}
.stDownloadButton > button:hover {{ background: {C['verde_medio']} !important; }}

/* ── Multiselect / selectbox ── */
span[data-baseweb="tag"] {{
    background: {C['azul_medio']} !important;
    color: white !important;
    border-radius: 4px !important;
}}
span[data-baseweb="tag"] span {{ color: white !important; font-size: 0.75rem !important; }}
/* ── Tarjeta de error ── */
.error-card {{
    background: #fff5f5;
    border: 1.5px solid #fca5a5;
    border-left: 5px solid #dc2626;
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin: 1rem 0;
}}
.error-card .error-title {{
    font-family: 'Montserrat', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: #991b1b;
    margin-bottom: 0.4rem;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}}
.error-card .error-body {{
    font-size: 0.83rem;
    color: #7f1d1d;
    line-height: 1.6;
    margin-bottom: 0.8rem;
}}
.error-card .error-fix {{
    background: #fef2f2;
    border-radius: 6px;
    padding: 0.6rem 0.9rem;
    font-size: 0.8rem;
    color: #991b1b;
    border: 1px solid #fca5a5;
}}
.error-card .error-fix strong {{
    display: block;
    margin-bottom: 0.3rem;
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: #dc2626;
}}
.error-cols {{
    display: flex;
    flex-wrap: wrap;
    gap: 0.4rem;
    margin-top: 0.5rem;
}}
.col-missing {{
    background: #fee2e2;
    color: #991b1b;
    font-family: 'Montserrat', monospace;
    font-size: 0.73rem;
    font-weight: 700;
    padding: 3px 10px;
    border-radius: 4px;
    border: 1px solid #fca5a5;
}}
.col-wrong-type {{
    background: #fff7ed;
    color: #9a3412;
    font-family: 'Montserrat', monospace;
    font-size: 0.73rem;
    font-weight: 700;
    padding: 3px 10px;
    border-radius: 4px;
    border: 1px solid #fed7aa;
}}
/* Tabla de referencia de columnas */
.ref-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.78rem;
    margin-top: 0.8rem;
    border-radius: 8px;
    overflow: hidden;
}}
.ref-table th {{
    background: #f1f5f9;
    color: {C['azul_oscuro']};
    font-weight: 700;
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    padding: 0.45rem 0.75rem;
    text-align: left;
    border-bottom: 2px solid {C['border']};
}}
.ref-table td {{
    padding: 0.4rem 0.75rem;
    border-bottom: 1px solid {C['border']};
    color: {C['text']};
}}
.ref-table tr:last-child td {{ border-bottom: none; }}
.ref-table code {{
    background: #f1f5f9;
    padding: 1px 6px;
    border-radius: 3px;
    font-size: 0.72rem;
    color: {C['azul_medio']};
}}
/* ── Tooltip de cálculo de días ── */
.dias-tip-wrap {{
    position: relative;
    display: inline-block;
}}
.dias-val-link {{
    font-family: 'DM Mono', monospace;
    font-weight: 600;
    font-size: 0.8rem;
    border-bottom: 1px dashed {C['azul_medio']};
    cursor: help;
    padding-bottom: 1px;
}}
.dias-tip-box {{
    display: none;
    position: fixed;
    background: {C['azul_oscuro']};
    color: #ffffff;
    border-radius: 8px;
    padding: 0.75rem 0.9rem;
    width: 255px;
    box-shadow: 0 8px 28px rgba(0,20,60,0.35);
    z-index: 99999;
    pointer-events: none;
    font-size: 0.74rem;
    line-height: 1.5;
}}
/* Flecha — clase aplicada por JS */
.dias-tip-box::after {{
    content: '';
    position: absolute;
    left: 50%;
    transform: translateX(-50%);
    border: 6px solid transparent;
}}
.dias-tip-box.tip-abajo::after {{
    bottom: 100%;
    border-bottom-color: {C['azul_oscuro']};
}}
.dias-tip-box.tip-arriba::after {{
    top: 100%;
    border-top-color: {C['azul_oscuro']};
}}
.dias-tip-title {{
    font-family: 'Montserrat', sans-serif;
    font-size: 0.62rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: {C['cian']};
    margin-bottom: 0.5rem;
}}
.dias-tip-row {{
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    gap: 0.5rem;
    margin: 0.15rem 0;
}}
.dias-tip-lbl {{
    color: rgba(255,255,255,0.6);
    font-size: 0.7rem;
    flex-shrink: 0;
}}
.dias-tip-val {{
    font-family: 'DM Mono', monospace;
    font-size: 0.72rem;
    font-weight: 600;
    color: #ffffff;
    text-align: right;
}}
.dias-tip-op {{
    font-size: 0.65rem;
    color: rgba(255,255,255,0.4);
    margin: 0.1rem 0;
    padding-left: 0.2rem;
}}
.dias-tip-sep {{
    border-top: 1px solid rgba(255,255,255,0.15);
    margin: 0.4rem 0 0.35rem;
}}
.dias-tip-result {{
    font-family: 'DM Mono', monospace;
    font-weight: 800;
    font-size: 0.88rem;
    color: {C['cian']};
    text-align: right;
}}
.dias-tip-nota {{
    font-size: 0.65rem;
    color: rgba(255,255,255,0.45);
    margin-top: 0.45rem;
    line-height: 1.4;
    border-top: 1px solid rgba(255,255,255,0.08);
    padding-top: 0.35rem;
}}
</style>
""", unsafe_allow_html=True)

# JS para tooltip dinámico — inyectado con components.v1.html para que se ejecute
import streamlit.components.v1 as components
components.html("""
<script>
(function() {
  var doc = window.parent.document;
  var win = window.parent;

  // ── Toggle contratos — via event delegation, no función global ────────────
  function initToggleCtto() {
    doc.querySelectorAll('.ctto-toggle').forEach(function(btn) {
      if (btn._cttoInit) return;
      btn._cttoInit = true;
      btn.addEventListener('click', function() {
        var id  = btn.getAttribute('data-target');
        var row = doc.getElementById(id);
        if (!row) return;
        var open = row.classList.toggle('visible');
        btn.classList.toggle('open', open);
      });
    });
  }

  function positionTip(trigger, tip, tipH, tipW) {
    var rect = trigger.getBoundingClientRect();
    var margin = 10;
    var spaceBelow = win.innerHeight - rect.bottom;
    var spaceAbove = rect.top;

    tip.style.display = 'block';
    tip.classList.remove('tip-abajo', 'tip-arriba');

    if (spaceBelow >= tipH + margin || spaceBelow >= spaceAbove) {
      tip.style.top    = (rect.bottom + 8) + 'px';
      tip.style.bottom = 'auto';
      tip.classList.add('tip-abajo');
    } else {
      tip.style.top    = (rect.top - tipH - 8) + 'px';
      tip.style.bottom = 'auto';
      tip.classList.add('tip-arriba');
    }

    var left = rect.left + rect.width / 2 - tipW / 2;
    left = Math.max(8, Math.min(left, win.innerWidth - tipW - 8));
    tip.style.left = left + 'px';
  }

  function initTooltips() {
    doc.querySelectorAll('.dias-tip-wrap').forEach(function(wrap) {
      if (wrap._tipInit) return;
      wrap._tipInit = true;
      var tip = wrap.querySelector('.dias-tip-box');
      if (!tip) return;
      wrap.addEventListener('mouseenter', function() { positionTip(wrap, tip, 220, 255); });
      wrap.addEventListener('mouseleave', function() { tip.style.display = 'none'; });
    });

    doc.querySelectorAll('.badge').forEach(function(badge) {
      if (badge._tipInit) return;
      badge._tipInit = true;
      var tip = badge.querySelector('.badge-tooltip');
      if (!tip) return;
      badge.addEventListener('mouseenter', function() { positionTip(badge, tip, 110, 240); });
      badge.addEventListener('mouseleave', function() { tip.style.display = 'none'; });
    });
  }

  function initAll() {
    initToggleCtto();
    initTooltips();
  }

  var observer = new MutationObserver(function() { initAll(); });
  observer.observe(doc.body, { childList: true, subtree: true });
  initAll();
})();
</script>
""", height=0, scrolling=False)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES DE VALIDACIÓN — deben ir ANTES de procesar()
# ─────────────────────────────────────────────────────────────────────────────
TABLA_ESPERADA = "MatrizSeguimientoEvaluacion"

COLUMNAS_ESPERADAS = {
    "ENTIDAD O SECRETARIA":                 ("texto",  [pl.Utf8, pl.String]),
    "BPIN":                                 ("texto",  [pl.Utf8, pl.String]),
    "NOMBRE PROYECTO":                      ("texto",  [pl.Utf8, pl.String]),
    "ESTADO PROYECTO":                      ("texto",  [pl.Utf8, pl.String]),
    "ESTADO CONTRATO":                      ("texto",  [pl.Utf8, pl.String]),
    "CPI":                                  ("número", [pl.Float32, pl.Float64, pl.Int32, pl.Int64]),
    "SPI":                                  ("número", [pl.Float32, pl.Float64, pl.Int32, pl.Int64]),
    "FECHA APROBACIÓN PROYECTO":            ("fecha",  [pl.Date, pl.Datetime]),
    "FECHA DE APERTURA DEL PRIMER PROCESO": ("fecha",  [pl.Date, pl.Datetime]),
    "FECHA SUSCRIPCION":                    ("fecha",  [pl.Date, pl.Datetime]),
    "FECHA ACTA INICIO":                    ("fecha",  [pl.Date, pl.Datetime]),
    "HORIZONTE DEL PROYECTO":               ("fecha",  [pl.Date, pl.Datetime]),
    "FECHA DE FINALIZACIÓN":                ("fecha",  [pl.Date, pl.Datetime]),
    "FECHA DE CORTE GESPROY":               ("fecha",  [pl.Date, pl.Datetime]),
}

TIPO_LABEL = {
    "texto":  "Texto",
    "número": "Número decimal",
    "fecha":  "Fecha",
}

TIPO_EJEMPLO = {
    "texto":  "Ej: «Infraestructura», «SIN CONTRATAR»",
    "número": "Ej: 0, 1.5, 0.87  (sin letras ni símbolos)",
    "fecha":  "Ej: 15/03/2024  (formato fecha de Excel)",
}

# ─────────────────────────────────────────────────────────────────────────────
# LÓGICA
# ─────────────────────────────────────────────────────────────────────────────
INTERVALOS = {
    "hito_1_val": [("0-100", 0, 100), ("101-150", 101, 150), ("151-180", 151, 180), (">180", 181, None)],
    "hito_2_val": [("0-100", 0, 100), ("101-150", 101, 150), ("151-180", 151, 180), (">180", 181, None)],
    "hito_3_val": [("0-30",  0,  30), ("31-45",   31,  45), ("46-60",   46,  60),  (">60",  61, None)],
    "hito_5_val": [("0-100", 0, 100), ("101-150", 101, 150), ("151-180", 151, 180), (">180", 181, None)],
}

# Semáforos: colores y mensajes por hito e intervalo
SEMAFOROS = {
    "hito_1_val": {
        "0-100":   ("green",  "Verde",   "Proyecto dentro de los tiempos para su primera apertura del proceso de contratación."),
        "101-150": ("yellow", "Naranja", "Proyecto en alerta: más de 100 días sin apertura del primer proceso precontractual."),
        "151-180": ("orange", "Rojo",    "Proyecto en alerta roja: más de 150 días sin apertura del primer proceso precontractual."),
        ">180":    ("black",  "Negro",   "Proyecto en alerta negra: más de 180 días sin apertura del primer proceso precontractual."),
    },
    "hito_2_val": {
        "0-100":   ("green",  "Verde",   "Proyecto dentro de los tiempos para la firma del primer contrato."),
        "101-150": ("yellow", "Naranja", "Proyecto en alerta: más de 100 días sin firma del primer contrato."),
        "151-180": ("orange", "Rojo",    "Proyecto en alerta roja: más de 150 días sin firma del primer contrato."),
        ">180":    ("black",  "Negro",   "Proyecto en alerta negra: más de 180 días sin firma del primer contrato."),
    },
    "hito_3_val": {
        "0-30":  ("green",  "Verde",   "Proyecto dentro de los tiempos para la firma del acta de inicio."),
        "31-45": ("yellow", "Naranja", "Proyecto en alerta: más de 30 días sin firma del acta de inicio."),
        "46-60": ("orange", "Rojo",    "Proyecto en alerta roja: más de 45 días sin firma del acta de inicio."),
        ">60":   ("black",  "Negro",   "Proyecto en alerta negra: más de 60 días sin firma del acta de inicio."),
    },
    "hito_4_val": {
        "0-1":   ("green",  "Verde",   "Proyecto presenta horizonte vigente."),
        "1.1-3": ("yellow", "Naranja", "Proyecto con horizonte vencido entre 1 y 3 meses."),
        "3.1-6": ("orange", "Rojo",    "Proyecto con horizonte vencido mayor a 3 meses."),
        ">6":    ("black",  "Negro",   "Proyecto con horizonte vencido mayor a 6 meses."),
    },
    "hito_5_val": {
        "0-100":   ("green",  "Verde",   "Proyecto dentro de los tiempos para pasar a estado 'Para cierre'."),
        "101-150": ("yellow", "Naranja", "Proyecto en alerta: más de 100 días desde su terminación sin pasar a 'Para cierre'."),
        "151-180": ("orange", "Rojo",    "Proyecto en alerta roja: más de 150 días desde su terminación."),
        ">180":    ("black",  "Negro",   "Proyecto en alerta negra: más de 180 días desde su terminación."),
    },
}

TABLA_DESCENTRALIZADAS = "OtrosEjecutoresDescentralizadas"
COLS_EVAL = [
    "CALIFICACIÓN DESEMPEÑO EN LA CONTRATACIÓN",
    "CALIFICACIÓN INFORMACIÓN A TIEMPO",
    "CALIFICACIÓN EJECUCIÓN DEL PROYECTO",
    "CALIFICACIÓN CALIDAD INFORMACIÓN",
]
COLS_EVAL_LABELS = [
    "Desempeño en contratación",
    "Información a tiempo",
    "Ejecución del proyecto",
    "Calidad de la información",
]

def clasificar_hito4_meses(col):
    """Hito 4 se clasifica en meses (días / 30), no en días directos."""
    meses = pl.col(col) / 30.0
    return (
        pl.when(pl.col(col).is_null()).then(None)
        .when(meses <= 1).then(pl.lit("0-1"))
        .when(meses <= 3).then(pl.lit("1.1-3"))
        .when(meses <= 6).then(pl.lit("3.1-6"))
        .otherwise(pl.lit(">6"))
    )

def clasificar(col, intervalos):
    expr = pl.when(pl.col(col).is_null()).then(None)
    for label, lo, hi in intervalos:
        cond = (pl.col(col) >= lo) & (pl.col(col) <= hi) if hi is not None else (pl.col(col) >= lo)
        expr = expr.when(cond).then(pl.lit(label))
    return expr.otherwise(None)

@st.cache_data
def procesar(file_bytes):
    df = pl.read_excel(io.BytesIO(file_bytes), table_name=TABLA_ESPERADA)

    DATE_COLS = [
        "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
        "FECHA SUSCRIPCION", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
        "FECHA DE FINALIZACIÓN", "FECHA DE CORTE GESPROY",
    ]

    # Cast robusto: maneja pl.Date nativo, Int32/Int64 (serial Excel = días desde 1899-12-30),
    # y Utf8/String (texto "DD/MM/YYYY" o "YYYY-MM-DD") que puede llegar desde GitHub.
    EXCEL_EPOCH = date(1899, 12, 30)
    cast_exprs = []
    for col in DATE_COLS:
        dtype = df[col].dtype
        if dtype == pl.Date:
            # Ya es fecha — no tocar
            cast_exprs.append(pl.col(col))
        elif dtype in (pl.Int32, pl.Int64, pl.UInt32, pl.UInt16):
            # Serial numérico de Excel → convertir a Date sumando días desde epoch
            cast_exprs.append(
                (pl.lit(EXCEL_EPOCH) + pl.duration(days=pl.col(col).cast(pl.Int64)))
                .cast(pl.Date)
                .alias(col)
            )
        elif dtype in (pl.Utf8, pl.String):
            # Texto — intentar múltiples formatos comunes
            cast_exprs.append(
                pl.coalesce([
                    pl.col(col).str.to_date("%d/%m/%Y", strict=False),
                    pl.col(col).str.to_date("%Y-%m-%d", strict=False),
                    pl.col(col).str.to_date("%m/%d/%Y", strict=False),
                    pl.col(col).str.to_date("%d-%m-%Y", strict=False),
                ]).alias(col)
            )
        else:
            # Fallback genérico
            cast_exprs.append(pl.col(col).cast(pl.Date, strict=False))

    df = (
        df.select(
            "ENTIDAD O SECRETARIA", "BPIN", "NOMBRE PROYECTO",
            "ESTADO PROYECTO", "ESTADO CONTRATO",
            "CPI", "SPI",
            *DATE_COLS,
        )
        .with_columns(cast_exprs)
        .with_columns(
            # Hito 1
            pl.when(
                ((pl.col("ESTADO PROYECTO") == "SIN CONTRATAR") | pl.col("ESTADO PROYECTO").is_null() | (pl.col("ESTADO PROYECTO") == "")) &
                (~pl.col("FECHA APROBACIÓN PROYECTO").is_null()) & (~pl.col("FECHA DE CORTE GESPROY").is_null())
            ).then((pl.col("FECHA DE CORTE GESPROY") - pl.col("FECHA APROBACIÓN PROYECTO")).dt.total_days()).otherwise(None).alias("hito_1_val"),
            # Hito 2 — clip(0) previene valores negativos por errores de datos
            # (ej: acta firmada antes de apertura del proceso)
            pl.when(
                ((pl.col("ESTADO PROYECTO") == "SIN CONTRATAR") | pl.col("ESTADO PROYECTO").is_null() | (pl.col("ESTADO PROYECTO") == "")) &
                (~pl.col("FECHA DE APERTURA DEL PRIMER PROCESO").is_null())
            ).then(
                (pl.col("FECHA ACTA INICIO") - pl.col("FECHA DE APERTURA DEL PRIMER PROCESO"))
                .dt.total_days().clip(lower_bound=0)
            ).otherwise(None).alias("hito_2_val"),
            # Hito 3
            pl.when(
                (pl.col("ESTADO PROYECTO") == "CONTRATADO SIN ACTA DE INICIO") &
                (~pl.col("FECHA SUSCRIPCION").is_null())
            ).then((pl.col("FECHA DE CORTE GESPROY") - pl.col("FECHA SUSCRIPCION")).dt.total_days()).otherwise(None).alias("hito_3_val"),
            # Hito 4
            pl.when(
                (pl.col("ESTADO PROYECTO") == "CONTRATADO EN EJECUCIÓN") &
                (pl.col("CPI") == 0) & (pl.col("SPI") == 0) &
                (pl.col("HORIZONTE DEL PROYECTO") <= pl.col("FECHA DE CORTE GESPROY"))
            ).then((pl.col("FECHA DE CORTE GESPROY") - pl.col("HORIZONTE DEL PROYECTO")).dt.total_days()).otherwise(None).alias("hito_4_val"),
            # Hito 5
            pl.when(~pl.col("FECHA DE FINALIZACIÓN").is_null()).then(
                (pl.col("FECHA DE CORTE GESPROY") - pl.col("FECHA DE FINALIZACIÓN")).dt.total_days()
            ).otherwise(None).alias("hito_5_val"),
            # Suspendidos — basado en ESTADO CONTRATO
            pl.when(
                pl.col("ESTADO CONTRATO").str.strip_chars().str.to_uppercase() == "SUSPENDIDO"
            ).then(pl.lit(1)).otherwise(None).alias("Suspendidos"),
            # Para cierre
            pl.when(pl.col("ESTADO PROYECTO") == "PARA CIERRE").then(pl.lit(1)).otherwise(None).alias("Para cierre"),
        )
        .with_columns(
            clasificar("hito_1_val", INTERVALOS["hito_1_val"]).alias("clasi_1"),
            clasificar("hito_2_val", INTERVALOS["hito_2_val"]).alias("clasi_2"),
            clasificar("hito_3_val", INTERVALOS["hito_3_val"]).alias("clasi_3"),
            clasificar_hito4_meses("hito_4_val").alias("clasi_4"),
            clasificar("hito_5_val", INTERVALOS["hito_5_val"]).alias("clasi_5"),
        )
    )
    return df

def _validar_cols_eval(df, cols, col_agrup):
    """
    Intenta castear las columnas de calificación a Float64.
    Retorna (df_casteado, errores) donde errores es lista de dicts con info de cada columna problemática.
    """
    errores = []
    cols_ok  = []
    for c in cols:
        if c not in df.columns:
            continue
        dtype_actual = str(df[c].dtype)
        # Verificar si ya es numérico
        if df[c].dtype in (pl.Float32, pl.Float64, pl.Int32, pl.Int64, pl.Int16, pl.UInt32):
            cols_ok.append(c)
            continue
        # Intentar cast — contar cuántos valores se perderían
        casteada = df[c].cast(pl.Float64, strict=False)
        nulos_antes  = df[c].is_null().sum()
        nulos_despues = casteada.is_null().sum()
        perdidos = int(nulos_despues - nulos_antes)
        if perdidos > 0:
            # Mostrar hasta 3 ejemplos de valores problemáticos
            ejemplos = (
                df.filter(df[c].is_not_null() & casteada.is_null())[c]
                .head(3).to_list()
            )
            errores.append({
                "col":        c,
                "tipo":       dtype_actual,
                "perdidos":   perdidos,
                "total":      df.height,
                "ejemplos":   ejemplos,
            })
        else:
            cols_ok.append(c)

    if cols_ok:
        df = df.with_columns([
            pl.col(c).cast(pl.Float64, strict=False) for c in cols_ok
        ])
    return df, cols_ok, errores


def _render_eval_errors(errores, contexto=""):
    """Muestra tarjetas de error amigables para columnas de calificación no numéricas."""
    tipo_amigable = {
        "string": "texto", "utf8": "texto", "str": "texto",
        "bool": "verdadero/falso", "date": "fecha", "datetime": "fecha",
    }
    st.warning(
        f"⚠️ Algunas calificaciones{' de ' + contexto if contexto else ''} "
        f"tienen datos que no pudieron leerse como números. "
        f"Esas columnas se excluyeron del cálculo.",
        icon=None,
    )
    for e in errores:
        pct = round(e["perdidos"] / e["total"] * 100, 1) if e["total"] else 0
        # Convertir tipo técnico a lenguaje amigable
        tipo_raw = e["tipo"].lower()
        tipo_legible = next((v for k, v in tipo_amigable.items() if k in tipo_raw), "no numérico")
        # Formatear ejemplos
        ejemplos_str = " · ".join([f'<code>{v}</code>' for v in e["ejemplos"]])
        st.markdown(f"""
<div class="error-card">
  <div class="error-title">&#9888; Calificación con valores incorrectos</div>
  <div class="error-body">
    La columna <strong>{e['col']}</strong> contiene valores de tipo <strong>{tipo_legible}</strong>
    en lugar de números.<br>
    Se encontraron <strong>{e['perdidos']} registro(s) con problemas</strong>
    de un total de {e['total']} ({pct}%).<br>
    Ejemplos de valores problemáticos encontrados: {ejemplos_str}
  </div>
  <div class="error-fix">
    <strong>Cómo corregirlo en Excel</strong>
    Abre el archivo, busca la columna <strong>{e['col']}</strong> y revisa
    que cada celda contenga únicamente un número (por ejemplo: <code>3.5</code>, <code>4</code> o <code>2.75</code>).
    Reemplaza cualquier texto, guion, «N/A» o celda vacía con el valor numérico correspondiente.
    Guarda el archivo y vuelve a cargarlo aquí.
  </div>
</div>""", unsafe_allow_html=True)


@st.cache_data
def procesar_descentralizadas(file_bytes):
    """Lee la tabla de descentralizadas y calcula promedios de calificación por EJECUTOR."""
    try:
        df = pl.read_excel(io.BytesIO(file_bytes), table_name=TABLA_DESCENTRALIZADAS)
        cols_disponibles = [c for c in COLS_EVAL if c in df.columns]
        if not cols_disponibles or "EJECUTOR" not in df.columns:
            return None, [], []
        df, cols_ok, errores = _validar_cols_eval(df, cols_disponibles, "EJECUTOR")
        if not cols_ok:
            return None, [], errores
        agg_exprs = [pl.col(c).mean().round(2).alias(c) for c in cols_ok]
        resultado = df.group_by("EJECUTOR").agg(agg_exprs).sort("EJECUTOR")
        return resultado, cols_ok, errores
    except Exception:
        return None, [], []


@st.cache_data
def procesar_eval_sucre(file_bytes):
    """Calcula promedios de calificación por ENTIDAD O SECRETARIA (tabla Sucre)."""
    try:
        df = pl.read_excel(io.BytesIO(file_bytes), table_name=TABLA_ESPERADA)
        cols_disponibles = [c for c in COLS_EVAL if c in df.columns]
        if not cols_disponibles or "ENTIDAD O SECRETARIA" not in df.columns:
            return None, [], []
        df, cols_ok, errores = _validar_cols_eval(df, cols_disponibles, "ENTIDAD O SECRETARIA")
        if not cols_ok:
            return None, [], errores
        agg_exprs = [pl.col(c).mean().round(2).alias(c) for c in cols_ok]
        resultado = df.group_by("ENTIDAD O SECRETARIA").agg(agg_exprs).sort("ENTIDAD O SECRETARIA")
        return resultado, cols_ok, errores
    except Exception:
        return None, [], []


def badge_html(val, hito_key=None):
    """Genera badge con punto de color semáforo y tooltip con mensaje."""
    if val is None: return ""
    val_str = str(val)
    cls_map = {
        "0-100":   "badge-green",  "0-30":  "badge-green",  "0-1":   "badge-green",
        "101-150": "badge-yellow", "31-45": "badge-yellow", "1.1-3": "badge-yellow",
        "151-180": "badge-orange", "46-60": "badge-orange", "3.1-6": "badge-orange",
        ">180":    "badge-black",  ">60":   "badge-black",  ">6":    "badge-black",
    }
    cls = cls_map.get(val_str, "badge-yellow")

    tooltip_html = ""
    if hito_key and hito_key in SEMAFOROS and val_str in SEMAFOROS[hito_key]:
        _, color_nombre, mensaje = SEMAFOROS[hito_key][val_str]
        tooltip_html = f'<span class="badge-tooltip"><strong style="color:#47b1d5;display:block;margin-bottom:3px">● {color_nombre}</strong>{mensaje}</span>'

    return (
        f'<span class="badge {cls}">'
        f'<span class="badge-dot"></span>'
        f'{val_str}'
        f'{tooltip_html}'
        f'</span>'
    )

def generar_excel(df_f_full, df_agr, clasi_por_entidad_map):
    """
    Genera reporte Excel formateado.
    df_f_full  : pl.DataFrame con todas las columnas (df_f con fechas + hitos)
    df_agr     : pl.DataFrame agrupado por entidad
    clasi_por_entidad_map: dict {entidad -> {clasi_k -> valor_mas_frecuente}}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    # ── Paleta ────────────────────────────────────────────────────────────────
    AZUL_OSC  = "003D6C"
    VERDE_OSC = "005931"
    AZUL_MED  = "1754AB"
    BLANCO    = "FFFFFF"
    GRIS_ALT  = "F7FAFD"
    GRIS_BRD  = "D1D5DB"

    SEM_FILL = {
        "0-100":   ("D1FAE5","065F46"), "0-30":  ("D1FAE5","065F46"), "0-1":   ("D1FAE5","065F46"),
        "101-150": ("FEF3C7","92400E"), "31-45": ("FEF3C7","92400E"), "1.1-3": ("FEF3C7","92400E"),
        "151-180": ("FFEDD5","9A3412"), "46-60": ("FFEDD5","9A3412"), "3.1-6": ("FFEDD5","9A3412"),
        ">180":    ("1E293B","F1F5F9"), ">60":   ("1E293B","F1F5F9"), ">6":    ("1E293B","F1F5F9"),
    }
    SEM_NOMBRE = {
        "0-100":"Verde","0-30":"Verde","0-1":"Verde",
        "101-150":"Naranja","31-45":"Naranja","1.1-3":"Naranja",
        "151-180":"Rojo","46-60":"Rojo","3.1-6":"Rojo",
        ">180":"Negro",">60":"Negro",">6":"Negro",
    }
    # Mensajes completos desde SEMAFOROS global
    SEM_MSG = {}
    for hk, vals in SEMAFOROS.items():
        for label, (_, color_nombre, msg) in vals.items():
            SEM_MSG[label] = msg

    def _side():  return Side(style="thin", color=GRIS_BRD)
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    def _font(bold=False, color="1A2332", size=9, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    def _fill(color): return PatternFill("solid", fgColor=color)
    def _align(h="left", wrap=True, v="top"):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _header_cell(cell, text):
        cell.value = text
        cell.font     = Font(name="Arial", bold=True, color=BLANCO, size=9)
        cell.fill     = _fill(AZUL_OSC)
        cell.alignment= Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border   = _border()

    def _data_cell(cell, value, bg=BLANCO, bold=False, color="1A2332", center=False, fmt=None):
        cell.value     = value
        cell.font      = _font(bold=bold, color=color)
        cell.fill      = _fill(bg)
        cell.alignment = _align("center" if center else "left")
        cell.border    = _border()
        if fmt: cell.number_format = fmt

    def _sem_cell(cell, clasi, bg_row=BLANCO):
        """Aplica color semáforo + comentario con mensaje."""
        clasi_s = str(clasi) if clasi and str(clasi) not in ("nan","None","") else None
        if clasi_s and clasi_s in SEM_FILL:
            bg, fg = SEM_FILL[clasi_s]
            cell.value     = clasi_s
            cell.font      = Font(name="Arial", bold=True, color=fg, size=9)
            cell.fill      = _fill(bg)
            cell.alignment = _align("center")
            cell.border    = _border()
            msg = SEM_MSG.get(clasi_s,"")
            nombre = SEM_NOMBRE.get(clasi_s,"")
            if msg:
                cell.comment = Comment(f"🔴 {nombre}\n{msg}", "Sistema", height=60, width=300)
        else:
            _data_cell(cell, "—", bg=bg_row, center=True, color="9CA3AF")

    def _title_row(ws, text, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text)
        c.font = Font(name="Arial", bold=True, size=13, color=BLANCO)
        c.fill = _fill(AZUL_OSC)
        c.alignment = _align("left")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c2 = ws.cell(2, 1, sub)
        c2.font = Font(name="Arial", size=9, color=BLANCO, italic=True)
        c2.fill = _fill(VERDE_OSC)
        c2.alignment = _align("left")
        ws.row_dimensions[2].height = 16

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1 · Resumen por Entidad
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen por Entidad"
    ws1.sheet_view.showGridLines = False

    H1_COLS = [
        ("Entidad / Secretaría",        28),
        ("H1 · Sin contratar\nSin apertura\nPromedio días", 13),
        ("H1\nAlerta",                   9),
        ("H2 · Sin contratar\nCon apertura\nPromedio días", 13),
        ("H2\nAlerta",                   9),
        ("H3 · Contratado\nSin acta inicio\nPromedio días", 13),
        ("H3\nAlerta",                   9),
        ("H4 · En ejecución\nRezagado\nPromedio meses",    13),
        ("H4\nAlerta",                   9),
        ("H5 · Terminados\nPromedio días",                 13),
        ("H5\nAlerta",                   9),
        ("Suspendidos",                 12),
        ("Para cierre",                 12),
        ("Total\nproyectos",            10),
    ]
    NCOLS1 = len(H1_COLS)
    _title_row(ws1,
               "Seguimiento y Evaluación · Regalías — Resumen por Entidad",
               f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Promedio de días/meses por hito   ·   Los colores indican el nivel de alerta",
               NCOLS1)

    # Fila de grupo de cabeceras (fila 3 = subgrupos)
    ws1.row_dimensions[3].height = 6   # separador visual

    # Headers (fila 4)
    for ci, (label, width) in enumerate(H1_COLS, 1):
        _header_cell(ws1.cell(4, ci), label)
        ws1.column_dimensions[get_column_letter(ci)].width = width
    ws1.row_dimensions[4].height = 42

    agr_pd = df_agr.to_pandas()
    # Columnas del agrupacion en orden garantizado
    AGR_COLS = list(agr_pd.columns)  # ['ENTIDAD O SECRETARIA', 'Hito 1 (días)', ..., 'Total']

    HITO_AGR = [
        ("Hito 1 (días)",  "clasi_1"),
        ("Hito 2 (días)",  "clasi_2"),
        ("Hito 3 (días)",  "clasi_3"),
        ("Hito 4 (días)",  "clasi_4"),
        ("Hito 5 (días)",  "clasi_5"),
    ]

    for ri, row_vals in enumerate(agr_pd.values.tolist(), 5):
        row_dict = dict(zip(AGR_COLS, row_vals))
        bg = GRIS_ALT if ri % 2 == 0 else BLANCO
        entidad = row_dict.get("ENTIDAD O SECRETARIA") or ""
        ws1.row_dimensions[ri].height = 18

        _data_cell(ws1.cell(ri, 1), entidad, bg=bg, bold=True, color=AZUL_MED)

        col = 2
        for dias_col, clasi_key in HITO_AGR:
            dias = row_dict.get(dias_col)
            _data_cell(ws1.cell(ri, col),
                       round(float(dias), 1) if dias is not None and str(dias) != "nan" else None,
                       bg=bg, center=True, fmt="#,##0.0")
            col += 1

            clasi = (clasi_por_entidad_map.get(entidad) or {}).get(clasi_key)
            _sem_cell(ws1.cell(ri, col), clasi, bg_row=bg)
            col += 1

        for extra_col, bold_it, extra_bg in [
            ("Suspendidos", False, None),
            ("Para cierre",  False, None),
            ("Total",        True,  "EFF6FF"),
        ]:
            val = row_dict.get(extra_col)
            v   = int(val) if val is not None and str(val) != "nan" else 0
            c   = ws1.cell(ri, col)
            _data_cell(c, v, bg=extra_bg or bg, bold=bold_it,
                       color=AZUL_MED if extra_bg else "1A2332", center=True)
            col += 1

    ws1.freeze_panes = ws1.cell(5, 2)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2 · Detalle por Proyecto
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle Proyectos")
    ws2.sheet_view.showGridLines = False

    # Columnas: info básica + fechas + hitos (días + alerta + mensaje)
    H2_COLS = [
        # (header, ancho, attr_en_df)
        ("Entidad /\nSecretaría",              26, "ENTIDAD O SECRETARIA"),
        ("BPIN",                               13, "BPIN"),
        ("Nombre del proyecto",                42, "NOMBRE PROYECTO"),
        ("Estado\nproyecto",                   18, "ESTADO PROYECTO"),
        ("Estado\ncontrato",                   18, "ESTADO CONTRATO"),
        # Fechas usadas para calcular hitos
        ("Fecha\naprobación",                  13, "FECHA APROBACIÓN PROYECTO"),
        ("Fecha apertura\nprimer proceso",      13, "FECHA DE APERTURA DEL PRIMER PROCESO"),
        ("Fecha\nsuscripción",                 13, "FECHA SUSCRIPCION"),
        ("Fecha acta\ninicio",                 13, "FECHA ACTA INICIO"),
        ("Horizonte\nproyecto",                13, "HORIZONTE DEL PROYECTO"),
        ("Fecha\nfinalización",                13, "FECHA DE FINALIZACIÓN"),
        ("Fecha corte\nGESPROY",               13, "FECHA DE CORTE GESPROY"),
        # Hito 1
        ("H1\ndías",                            8, "hito_1_val"),
        ("H1\nalerta",                          9, "clasi_1"),
        ("H1 · Mensaje de alerta",             34, "_msg_1"),
        # Hito 2
        ("H2\ndías",                            8, "hito_2_val"),
        ("H2\nalerta",                          9, "clasi_2"),
        ("H2 · Mensaje de alerta",             34, "_msg_2"),
        # Hito 3
        ("H3\ndías",                            8, "hito_3_val"),
        ("H3\nalerta",                          9, "clasi_3"),
        ("H3 · Mensaje de alerta",             34, "_msg_3"),
        # Hito 4
        ("H4\ndías",                            8, "hito_4_val"),
        ("H4\nalerta",                          9, "clasi_4"),
        ("H4 · Mensaje de alerta",             34, "_msg_4"),
        # Hito 5
        ("H5\ndías",                            8, "hito_5_val"),
        ("H5\nalerta",                          9, "clasi_5"),
        ("H5 · Mensaje de alerta",             34, "_msg_5"),
        # Flags
        ("Suspendido",                         11, "Suspendidos"),
        ("Para\ncierre",                       10, "Para cierre"),
    ]
    NCOLS2 = len(H2_COLS)
    _title_row(ws2,
               "Seguimiento y Evaluación · Regalías — Detalle por Proyecto",
               f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Incluye fechas usadas en el cálculo de hitos y niveles de alerta con mensajes",
               NCOLS2)

    ws2.row_dimensions[3].height = 6  # separador

    for ci, (label, width, _) in enumerate(H2_COLS, 1):
        _header_cell(ws2.cell(3, ci), label)
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.row_dimensions[3].height = 42

    DATE_COLS = {
        "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
        "FECHA SUSCRIPCION", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
        "FECHA DE FINALIZACIÓN", "FECHA DE CORTE GESPROY",
    }
    HITO_CLASI_PAIRS = [
        ("hito_1_val","clasi_1"),("hito_2_val","clasi_2"),
        ("hito_3_val","clasi_3"),("hito_4_val","clasi_4"),
        ("hito_5_val","clasi_5"),
    ]
    NUM_COLS = {"hito_1_val","hito_2_val","hito_3_val","hito_4_val","hito_5_val"}
    FLAG_COLS = {"Suspendidos","Para cierre"}

    rows = df_f_full.to_dicts()
    for ri2, row in enumerate(rows, 4):
        bg = GRIS_ALT if ri2 % 2 == 0 else BLANCO
        ws2.row_dimensions[ri2].height = 80

        # Pre-calcular mensajes por hito
        msgs = {}
        for n, (_, clasi_col) in enumerate(HITO_CLASI_PAIRS, 1):
            cv = row.get(clasi_col)
            cv_s = str(cv) if cv and str(cv) not in ("nan","None","") else None
            msgs[f"_msg_{n}"] = SEM_MSG.get(cv_s, "") if cv_s else ""

        for ci, (_, _, attr) in enumerate(H2_COLS, 1):
            cell = ws2.cell(ri2, ci)
            val  = row.get(attr) if not attr.startswith("_msg") else msgs.get(attr, "")

            if attr in ("clasi_1","clasi_2","clasi_3","clasi_4","clasi_5"):
                _sem_cell(cell, val, bg_row=bg)

            elif attr in NUM_COLS:
                v = round(float(val), 1) if val is not None and str(val) != "nan" else None
                _data_cell(cell, v, bg=bg, center=True,
                           color="1A2332", fmt="#,##0.0")

            elif attr in DATE_COLS:
                if val is not None and str(val) not in ("nan","NaT","None",""):
                    if isinstance(val, (_dt.date, _dt.datetime)):
                        cell.value          = val
                        cell.number_format  = "DD/MM/YYYY"
                        cell.font           = _font()
                        cell.fill           = _fill(bg)
                        cell.alignment      = _align("center")
                        cell.border         = _border()
                    else:
                        _data_cell(cell, str(val), bg=bg, center=True)
                else:
                    _data_cell(cell, "—", bg=bg, center=True, color="9CA3AF")

            elif attr in FLAG_COLS:
                es_si = val is not None and str(val) not in ("nan","None","0","") and val != 0
                _data_cell(cell, "Sí" if es_si else "No",
                           bg="FEF3C7" if es_si else bg,
                           center=True,
                           bold=es_si,
                           color="92400E" if es_si else "6B7280")

            elif attr.startswith("_msg"):
                _data_cell(cell, val, bg=bg, color="374151")
                cell.font = Font(name="Arial", size=8, color="374151", italic=True)

            else:
                _data_cell(cell, val if val else "—", bg=bg)

    ws2.freeze_panes = ws2.cell(4, 3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def th(label, titulo, desc):
    return f"""<th><div class="th-wrap">{label}<span class="th-icon">?</span>
    <div class="th-tooltip"><strong>{titulo}</strong>{desc}</div></div></th>"""

# ─────────────────────────────────────────────────────────────────────────────
# VALIDACIÓN
# ─────────────────────────────────────────────────────────────────────────────

def error_card(titulo, cuerpo, solucion):
    return f"""
    <div class="error-card">
        <div class="error-title">&#9888; {titulo}</div>
        <div class="error-body">{cuerpo}</div>
        <div class="error-fix"><strong>Cómo solucionarlo</strong>{solucion}</div>
    </div>"""

def _tipo_amigable(dtype_str):
    """Convierte tipo técnico de Polars a nombre comprensible."""
    d = dtype_str.lower()
    if "utf" in d or "str" in d or "cat" in d: return "Texto"
    if "float" in d or "int" in d:             return "Número"
    if "date" in d or "time" in d:             return "Fecha"
    if "bool" in d:                            return "Verdadero/Falso"
    return dtype_str

def validar_archivo(file_bytes):
    """Retorna (df, errores_html). Si hay errores, df es None."""
    errores = []

    # 1. Verificar que la tabla existe
    try:
        df_raw = pl.read_excel(io.BytesIO(file_bytes), table_name=TABLA_ESPERADA)
    except Exception as e:
        msg = str(e)
        if "table" in msg.lower() or "not found" in msg.lower() or "name" in msg.lower():
            errores.append(error_card(
                "Tabla no encontrada",
                f"No se encontró una tabla con el nombre <b>{TABLA_ESPERADA}</b> en el archivo. "
                f"Es posible que el nombre haya sido cambiado o que los datos no estén definidos como tabla de Excel.",
                f"En Excel, selecciona el rango de datos → <b>Insertar → Tabla</b>, y asegúrate de que el nombre "
                f"sea exactamente <code>{TABLA_ESPERADA}</code> (sin espacios adicionales, respetando mayúsculas)."
            ))
        else:
            errores.append(error_card(
                "Error al leer el archivo",
                f"El archivo no pudo ser leído correctamente.",
                "Verifica que el archivo no esté dañado, que tenga extensión <b>.xlsx</b> y que no esté abierto en Excel al momento de cargarlo."
            ))
        return None, errores

    cols_actuales = set(df_raw.columns)

    # 2. Columnas faltantes
    faltantes = [c for c in COLUMNAS_ESPERADAS if c not in cols_actuales]
    if faltantes:
        chips = "".join(f"<span class='col-missing'>{c}</span>" for c in faltantes)
        errores.append(error_card(
            f"{'Columna faltante' if len(faltantes) == 1 else f'{len(faltantes)} columnas faltantes'}",
            f"Las siguientes columnas no fueron encontradas en la tabla:<div class='error-cols'>{chips}</div>"
            f"<div style='margin-top:0.5rem;font-size:0.79rem;color:#7f1d1d'>Puede que el nombre haya sido modificado por error. "
            f"Compara con la lista de columnas esperadas al final de esta pantalla.</div>",
            "Abre el archivo en Excel, ve a la tabla <b>MatrizSeguimientoEvaluacion</b> y verifica que los encabezados "
            "coincidan exactamente (respeta mayúsculas, tildes y espacios). No renombres las columnas originales."
        ))

    # 3. Tipo de datos incorrecto — solo columnas que sí existen
    tipo_incorrecto = []
    for col, (tipo_label, tipos_validos) in COLUMNAS_ESPERADAS.items():
        if col not in cols_actuales:
            continue
        dtype = df_raw[col].dtype
        if tipo_label == "fecha":
            continue  # el cast strict=False lo maneja en procesar
        if dtype not in tipos_validos:
            tipo_incorrecto.append((col, tipo_label, str(dtype)))

    if tipo_incorrecto:
        chips = "".join(
            f"<span class='col-wrong-type'>{col}</span>"
            for col, _, _ in tipo_incorrecto
        )
        detalles = "".join(
            f"<li style='margin-bottom:4px'><b>{col}</b>: "
            f"el sistema encontró <b>{_tipo_amigable(dtype_actual)}</b>, "
            f"pero esperaba <b>{TIPO_LABEL[tipo_label]}</b> — {TIPO_EJEMPLO[tipo_label]}</li>"
            for col, tipo_label, dtype_actual in tipo_incorrecto
        )
        errores.append(error_card(
            f"{'Tipo de dato incorrecto' if len(tipo_incorrecto) == 1 else f'Tipo de dato incorrecto en {len(tipo_incorrecto)} columnas'}",
            f"Las siguientes columnas tienen un tipo de dato que no corresponde:<div class='error-cols'>{chips}</div>"
            f"<ul style='margin-top:0.6rem;font-size:0.8rem;padding-left:1.2rem'>{detalles}</ul>",
            "En Excel, selecciona la columna señalada y revisa el formato de las celdas (menú <b>Inicio → Número</b>). "
            "Las columnas <b>CPI</b> y <b>SPI</b> deben tener formato <b>Número</b>. "
            "Si los valores están alineados a la izquierda, es probable que estén guardados como texto: "
            "selecciona la columna → <b>Datos → Texto en columnas</b> → finalizar."
        ))

    if errores:
        return None, errores

    return df_raw, []

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — carga y filtros
# ─────────────────────────────────────────────────────────────────────────────

# URL del archivo por defecto en GitHub (rama main, carpeta data/)
# Reemplaza esta URL con la tuya antes de hacer deploy
GITHUB_RAW_URL          = "https://raw.githubusercontent.com/Dona121/Matriz-Evaluacion-Regalias/main/data/MatrizSeguimientoEvaluacion.xlsx"
GITHUB_CONTRATOS_URL    = "https://raw.githubusercontent.com/Dona121/Matriz-Evaluacion-Regalias/main/data/CG-cttos.xlsx"

@st.cache_data(show_spinner=False, ttl=3600)
def _cargar_desde_github(url: str):
    """Descarga el Excel desde GitHub Raw y devuelve los bytes. Cachea 1 hora."""
    try:
        import urllib.request
        with urllib.request.urlopen(url, timeout=15) as r:
            return r.read()
    except Exception:
        return None

def _parse_valor(s):
    """
    Convierte string de valor monetario a float.
    Maneja formatos COP: "1,234,567.89", "1.234.567,89", "1234567", etc.
    Definida a nivel de módulo para serialización eficiente en map_elements().
    """
    if s is None:
        return None
    s = str(s).strip().replace("$", "").replace(" ", "")
    if not s or s in ("", "None", "-", "—"):
        return None
    has_comma = "," in s
    has_dot   = "." in s
    try:
        if has_comma and has_dot:
            last_comma = s.rfind(",")
            last_dot   = s.rfind(".")
            if last_dot > last_comma:
                s = s.replace(",", "")
            else:
                s = s.replace(".", "").replace(",", ".")
        elif has_comma:
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) <= 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        elif has_dot:
            parts = s.split(".")
            if len(parts) > 2:
                s = s.replace(".", "")
        return float(s)
    except Exception:
        return None

@st.cache_data(show_spinner=False)
def procesar_contratos(file_bytes):
    """
    Lee y limpia el reporte de contratos de GESPROY.
    Retorna (df, diagnostico_str) — df puede ser None si falla.
    diagnostico tiene info útil para debugging.
    """
    COLS_CONTRATOS = [
        "BPIN",
        "NO. PROCESO PRECONTRACTUAL",
        "MODALIDAD CONTRATACION",
        "TIPO CONTRATO",
        "CONTRATO OBJETO",
        "CONTRATO VALOR TOTAL",
        "ESTADO CONTRATO",
    ]
    diag = []
    try:
        # ── Leer raw sin encabezado (GESPROY exporta 2 filas de meta antes de los datos) ──
        df_raw = pl.read_excel(
            io.BytesIO(file_bytes),
            has_header=False,
            infer_schema_length=0,  # todo como string inicialmente
        )
        diag.append(f"Filas raw: {df_raw.height}, Cols: {df_raw.width}")

        if df_raw.height < 3:
            return None, f"Archivo muy pequeño ({df_raw.height} filas)"

        # ── Detectar fila de encabezados robustamente ────────────────────────
        # Buscar la fila que contiene "BPIN" en alguna celda (puede ser fila 0 o 1)
        header_row_idx = None
        for row_idx in range(min(5, df_raw.height)):
            row_vals = [str(v).strip().upper() for v in df_raw.row(row_idx) if v is not None]
            if "BPIN" in row_vals:
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            diag.append("ERROR: no se encontró fila con 'BPIN'")
            diag.append(f"Fila 0: {list(df_raw.row(0))[:8]}")
            diag.append(f"Fila 1: {list(df_raw.row(1))[:8]}")
            return None, " | ".join(diag)

        diag.append(f"Fila de headers detectada: {header_row_idx}")

        # ── Construir encabezados desde la fila detectada ────────────────────
        encabezados_raw = df_raw.row(header_row_idx)
        encabezados = []
        seen = {}
        for i, v in enumerate(encabezados_raw):
            name = str(v).strip() if v is not None and str(v).strip() not in ("", "None") else f"_col_{i}"
            # Deduplicar nombres repetidos
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            encabezados.append(name)

        # ── Renombrar y saltar filas de encabezado ───────────────────────────
        df = (
            df_raw
            .rename(dict(zip(df_raw.columns, encabezados)))
            .slice(header_row_idx + 1)  # datos empiezan después del header
        )
        diag.append(f"Columnas detectadas (primeras 10): {list(df.columns[:10])}")
        diag.append(f"Filas de datos: {df.height}")

        # ── Verificar columnas necesarias ────────────────────────────────────
        cols_presentes = set(df.columns)
        faltantes = [c for c in COLS_CONTRATOS if c not in cols_presentes]
        if faltantes:
            diag.append(f"ERROR: columnas faltantes: {faltantes}")
            diag.append(f"Columnas disponibles: {sorted(cols_presentes)}")
            return None, " | ".join(diag)

        df = df.select(COLS_CONTRATOS)

        # ── Limpieza ─────────────────────────────────────────────────────────
        # 1. Strip todas las columnas texto
        str_cols = [c for c in COLS_CONTRATOS if c != "CONTRATO VALOR TOTAL"]
        df = df.with_columns([
            pl.col(c).cast(pl.Utf8, strict=False).str.strip_chars().alias(c)
            for c in str_cols
        ])

        # 2. BPIN: forzar string, quitar puntos/comas/espacios/guiones
        df = df.with_columns(
            pl.col("BPIN")
            .cast(pl.Utf8, strict=False)
            .str.strip_chars()
            .str.replace_all(r"[.\-,\s]", "")
            .alias("BPIN")
        )

        # 3. Filtrar filas sin BPIN válido
        df = df.filter(
            pl.col("BPIN").is_not_null() &
            (pl.col("BPIN").str.len_chars() >= 5) &
            (pl.col("BPIN") != "None") &
            (pl.col("BPIN") != "BPIN")  # filtrar si quedó alguna fila de encabezado
        )
        diag.append(f"Filas con BPIN válido: {df.height}")

        # 4. Muestra de BPINs para diagnóstico
        bpins_muestra = df["BPIN"].head(5).to_list()
        diag.append(f"BPINs muestra: {bpins_muestra}")

        # 5. CONTRATO VALOR TOTAL → Float64 (usa _parse_valor definida a nivel de módulo)
        df = df.with_columns(
            pl.col("CONTRATO VALOR TOTAL")
            .cast(pl.Utf8, strict=False)
            .map_elements(_parse_valor, return_dtype=pl.Float64)
            .alias("CONTRATO VALOR TOTAL")
        )

        # 6. Deduplicar y uppercase estado
        df = df.unique()
        df = df.with_columns(
            pl.col("ESTADO CONTRATO")
            .cast(pl.Utf8, strict=False)
            .str.strip_chars()
            .str.to_uppercase()
            .alias("ESTADO CONTRATO")
        )

        return (df if df.height > 0 else None), " | ".join(diag)

    except Exception as e:
        diag.append(f"EXCEPCIÓN: {type(e).__name__}: {e}")
        return None, " | ".join(diag)

with st.sidebar:
    st.markdown("<div class='sidebar-section'>📁 Datos</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Subir otro archivo Excel", type=["xlsx"], label_visibility="collapsed")
    if uploaded:
        st.success("Usando el archivo subido manualmente.")
    else:
        st.markdown(
            f"<p style='font-size:0.7rem;color:rgba(255,255,255,0.5);margin:0.3rem 0 0'>"
            f"Cargando datos desde el repositorio por defecto. "
            f"Sube un archivo para usar datos distintos.</p>",
            unsafe_allow_html=True,
        )
    st.markdown("<div class='sidebar-section'>📋 Contratos</div>", unsafe_allow_html=True)
    uploaded_cttos = st.file_uploader(
        "Archivo de contratos (CG-cttos)",
        type=["xlsx"],
        label_visibility="collapsed",
        key="uploader_contratos",
    )
    if uploaded_cttos:
        st.success("Usando contratos cargados manualmente.")
    else:
        st.markdown(
            f"<p style='font-size:0.7rem;color:rgba(255,255,255,0.5);margin:0.3rem 0 0'>"
            f"Contratos desde el repositorio por defecto.</p>",
            unsafe_allow_html=True,
        )

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="page-header">
  <div>
    <h1>Seguimiento y Evaluación · Regalías</h1>
    <p>Matriz de seguimiento de proyectos por hito de gestión</p>
  </div>
</div>
""", unsafe_allow_html=True)

# Determinar fuente de datos: manual > GitHub > error
# IMPORTANTE: UploadedFile.read() es un stream — se consume en el primer rerun.
# Cacheamos los bytes en session_state indexados por nombre+tamaño del archivo.
if uploaded is not None:
    _upload_id = f"{uploaded.name}_{uploaded.size}"
    if st.session_state.get("_upload_id") != _upload_id:
        st.session_state["_upload_id"]    = _upload_id
        st.session_state["_file_bytes"]   = uploaded.read()
    file_bytes = st.session_state["_file_bytes"]
else:
    st.session_state.pop("_upload_id",  None)
    st.session_state.pop("_file_bytes", None)
    with st.spinner("Cargando datos desde el repositorio…"):
        file_bytes = _cargar_desde_github(GITHUB_RAW_URL)

if file_bytes is None:
    st.error(
        "No se pudo cargar el archivo de datos. "
        "Verifica que la URL en `GITHUB_RAW_URL` sea correcta y que el repositorio sea público, "
        "o sube el archivo manualmente desde el panel izquierdo."
    )
    st.stop()
df_raw, errores = validar_archivo(file_bytes)

if errores:
    muted = C["muted"]
    ref_rows = "".join(
        f"<tr><td><code>{col}</code></td>"
        f"<td><b>{TIPO_LABEL[tipo]}</b></td>"
        f"<td style='color:{muted}'>{TIPO_EJEMPLO[tipo]}</td></tr>"
        for col, (tipo, _) in COLUMNAS_ESPERADAS.items()
    )
    ref_table = f"""
    <table class="ref-table">
        <thead><tr>
            <th>Nombre exacto de la columna</th>
            <th>Tipo de dato esperado</th>
            <th>Ejemplo</th>
        </tr></thead>
        <tbody>{ref_rows}</tbody>
    </table>"""

    st.markdown(f"""
    <div style="background:{C['white']};border-radius:12px;padding:1.5rem 1.8rem;
                box-shadow:0 1px 6px rgba(0,0,0,0.07);margin-top:0.5rem">
        <div style="font-family:'Montserrat',sans-serif;font-size:1rem;font-weight:700;
                    color:{C['azul_oscuro']};margin-bottom:0.3rem">
            No se pudo cargar el archivo
        </div>
        <div style="font-size:0.83rem;color:{C['muted']};margin-bottom:1rem">
            Se encontraron los siguientes problemas. Corrígelos en Excel y vuelve a cargar el archivo.
        </div>
        {"".join(errores)}
        <details style="margin-top:1.2rem">
            <summary style="font-size:0.8rem;font-weight:600;color:{C['azul_medio']};
                            cursor:pointer;user-select:none;list-style:none;display:flex;align-items:center;gap:6px">
                &#9432; Ver referencia completa: tabla y columnas esperadas
            </summary>
            <div style="margin-top:0.8rem">
                <div style="font-size:0.78rem;font-weight:700;text-transform:uppercase;
                            letter-spacing:0.8px;color:{C['muted']};margin-bottom:0.3rem">
                    Nombre de la tabla en Excel
                </div>
                <div style="font-size:0.85rem;margin-bottom:1rem">
                    La tabla debe llamarse exactamente:
                    <code style="background:#f1f5f9;padding:2px 8px;border-radius:4px;
                    color:{C['azul_medio']};font-size:0.85rem">{TABLA_ESPERADA}</code>
                    <span style="font-size:0.78rem;color:{C['muted']}">
                    — puedes verificarlo en Excel seleccionando cualquier celda de la tabla
                    y mirando el cuadro de nombre en la esquina superior izquierda.</span>
                </div>
                <div style="font-size:0.78rem;font-weight:700;text-transform:uppercase;
                            letter-spacing:0.8px;color:{C['muted']};margin-bottom:0.3rem">
                    Columnas esperadas
                </div>
                <div style="font-size:0.78rem;color:{C['muted']};margin-bottom:0.6rem">
                    Los nombres deben ser exactamente iguales, incluyendo tildes, espacios y mayúsculas.
                </div>
                {ref_table}
            </div>
        </details>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

try:
    df = procesar(file_bytes)
except Exception as e:
    msg = str(e)
    # Detectar columna faltante en procesar (por si pasó la validación pero falla el select)
    col_hint = ""
    for col in COLUMNAS_ESPERADAS:
        if col.lower() in msg.lower():
            col_hint = f"<br>Columna relacionada: <code>{col}</code>"
            break
    st.markdown(f"""
    <div style="background:{C['white']};border-radius:12px;padding:1.5rem 1.8rem;
                box-shadow:0 1px 6px rgba(0,0,0,0.07);margin-top:0.5rem">
        <div style="font-family:'Montserrat',sans-serif;font-size:1rem;font-weight:700;
                    color:{C['azul_oscuro']};margin-bottom:0.8rem">
            Error al procesar el archivo
        </div>
        <div class="error-card">
            <div class="error-title">&#9888; Problema inesperado al leer los datos</div>
            <div class="error-body">
                El archivo fue cargado pero ocurrió un error al procesar su contenido.{col_hint}
            </div>
            <div class="error-fix">
                <strong>Cómo solucionarlo</strong>
                Verifica que ninguna columna haya sido renombrada o eliminada en la tabla
                <code>{TABLA_ESPERADA}</code>. Si el problema persiste, intenta exportar el archivo
                de nuevo desde el sistema origen y cargarlo sin modificaciones.
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CARGA DE CONTRATOS
# ─────────────────────────────────────────────────────────────────────────────
if uploaded_cttos is not None:
    _cttos_id = f"{uploaded_cttos.name}_{uploaded_cttos.size}"
    if st.session_state.get("_cttos_id") != _cttos_id:
        st.session_state["_cttos_id"]    = _cttos_id
        st.session_state["_cttos_bytes"] = uploaded_cttos.read()
    _cttos_bytes = st.session_state["_cttos_bytes"]
else:
    st.session_state.pop("_cttos_id",    None)
    st.session_state.pop("_cttos_bytes", None)
    with st.spinner("Cargando contratos desde el repositorio…"):
        _cttos_bytes = _cargar_desde_github(GITHUB_CONTRATOS_URL)

df_contratos, _cttos_diag = procesar_contratos(_cttos_bytes) if _cttos_bytes else (None, "No se obtuvieron bytes del archivo de contratos")

# ─────────────────────────────────────────────────────────────────────────────
# FILTROS EN SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
HITOS = {
    "H1 · Sin contratar sin apertura":    ("hito_1_val", "clasi_1"),
    "H2 · Sin contratar con apertura":    ("hito_2_val", "clasi_2"),
    "H3 · Contratado sin acta de inicio": ("hito_3_val", "clasi_3"),
    "H4 · En ejecución rezagado":         ("hito_4_val", "clasi_4"),
    "H5 · Proyectos terminados":          ("hito_5_val", "clasi_5"),
}

with st.sidebar:
    st.markdown("<div class='sidebar-section'>Filtros</div>", unsafe_allow_html=True)

    entidades   = sorted(df["ENTIDAD O SECRETARIA"].drop_nulls().unique().to_list())
    sel_entidades = st.multiselect("Entidad / Secretaría", entidades, default=entidades)

    estados_raw  = sorted(df["ESTADO PROYECTO"].drop_nulls().unique().to_list())
    tiene_sin    = df["ESTADO PROYECTO"].is_null().any() or (df["ESTADO PROYECTO"] == "").any()
    opciones_est = estados_raw + (["(Sin estado)"] if tiene_sin else [])
    sel_estados  = st.multiselect("Estado del proyecto", opciones_est, default=opciones_est)

# ─────────────────────────────────────────────────────────────────────────────
# FILTRAR
# ─────────────────────────────────────────────────────────────────────────────
filter_expr = pl.col("ENTIDAD O SECRETARIA").is_in(sel_entidades)
if sel_estados:
    incluir_sin   = "(Sin estado)" in sel_estados
    estados_reales = [e for e in sel_estados if e != "(Sin estado)"]
    if estados_reales and incluir_sin:
        estado_expr = pl.col("ESTADO PROYECTO").is_in(estados_reales) | pl.col("ESTADO PROYECTO").is_null() | (pl.col("ESTADO PROYECTO") == "")
    elif estados_reales:
        estado_expr = pl.col("ESTADO PROYECTO").is_in(estados_reales)
    elif incluir_sin:
        estado_expr = pl.col("ESTADO PROYECTO").is_null() | (pl.col("ESTADO PROYECTO") == "")
    else:
        estado_expr = pl.lit(False)
    filter_expr = filter_expr & estado_expr
df_f = df.filter(filter_expr)

# ─────────────────────────────────────────────────────────────────────────────
# KPIs — jerarquía visual
# ─────────────────────────────────────────────────────────────────────────────
total_proy      = df_f.height
total_entidades = df_f["ENTIDAD O SECRETARIA"].n_unique()
suspendidos     = int(df_f["Suspendidos"].drop_nulls().sum()) if df_f["Suspendidos"].drop_nulls().len() > 0 else 0
para_cierre     = int(df_f["Para cierre"].drop_nulls().sum()) if df_f["Para cierre"].drop_nulls().len() > 0 else 0

# Conteo por estado de proyecto
estados_conteo = (
    df_f.group_by("ESTADO PROYECTO")
    .agg(pl.len().alias("n"))
    .sort("n", descending=True)
)
estado_items = ""
for row_e in estados_conteo.to_dicts():
    est = row_e["ESTADO PROYECTO"] or "(Sin estado)"
    n   = row_e["n"]
    # Color de punto según estado
    eu = est.strip().upper()
    dot_colors = {
        "CONTRATADO EN EJECUCIÓN": C["verde_medio"],
        "TERMINADO":               C["muted"],
        "SIN CONTRATAR":           C["cian"],
        "PARA CIERRE":             C["cafe"],
        "CONTRATADO SIN ACTA DE INICIO": C["azul_medio"],
        "SUSPENDIDO":              C["naranja_osc"],
    }
    dot = dot_colors.get(eu, C["muted"])
    estado_items += (
        f'<div class="estado-kpi-row">'
        f'<span class="estado-kpi-label"><span style="display:inline-block;width:7px;height:7px;'
        f'border-radius:50%;background:{dot};margin-right:5px;flex-shrink:0"></span>{est}</span>'
        f'<span class="estado-kpi-n">{n}</span></div>'
    )

st.markdown("<div style='height:0.2rem'></div>", unsafe_allow_html=True)
ka, kb, kd = st.columns([1.3, 1.3, 3.2])

with ka:
    st.markdown(f"""
    <div class="kpi-main">
        <div class="label">Total proyectos</div>
        <div class="value">{total_proy}</div>
        <div class="sub">en los filtros activos</div>
    </div>""", unsafe_allow_html=True)

with kb:
    st.markdown(f"""
    <div class="kpi-main" style="background:{C['verde_oscuro']}">
        <div class="label">Entidades</div>
        <div class="value">{total_entidades}</div>
        <div class="sub">secretarías / dependencias</div>
    </div>""", unsafe_allow_html=True)

with kd:
    st.markdown(f"""
    <div class="kpi-estados">
        <div class="kpi-estados-title">Proyectos por estado</div>
        <div class="kpi-estados-grid">{estado_items}</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# AGRUPACIÓN
# ─────────────────────────────────────────────────────────────────────────────
agrupacion = (
    df_f.group_by("ENTIDAD O SECRETARIA")
    .agg(
        pl.col("hito_1_val").mean().round(1).alias("Hito 1 (días)"),
        pl.col("hito_2_val").mean().round(1).alias("Hito 2 (días)"),
        pl.col("hito_3_val").mean().round(1).alias("Hito 3 (días)"),
        pl.col("hito_4_val").mean().round(1).alias("Hito 4 (días)"),
        pl.col("hito_5_val").mean().round(1).alias("Hito 5 (días)"),
        pl.col("Suspendidos").sum().alias("Suspendidos"),
        pl.col("Para cierre").sum().alias("Para cierre"),
        pl.len().alias("Total"),
    )
    .sort("ENTIDAD O SECRETARIA")
)

# Clasificación modal por entidad — un solo group_by vectorizado en lugar de N filtros separados
# Equivalente a: para cada entidad, la clasificación más frecuente en cada hito
_CLASI_COLS = ["clasi_1", "clasi_2", "clasi_3", "clasi_4", "clasi_5"]

def _moda_str(s: pl.Series) -> str | None:
    """Devuelve el valor más frecuente de una serie string, o None si está vacía."""
    vals = s.drop_nulls()
    if len(vals) == 0:
        return None
    return vals.value_counts().sort("count", descending=True)[0, 0]

_clasi_modal = (
    df_f
    .group_by("ENTIDAD O SECRETARIA")
    .agg([pl.col(c).map_elements(_moda_str, return_dtype=pl.Utf8).first().alias(c)
          for c in _CLASI_COLS])
)

clasi_por_entidad = {
    row["ENTIDAD O SECRETARIA"]: {c: row[c] for c in _CLASI_COLS}
    for row in _clasi_modal.to_dicts()
}

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIONES Y CONSTANTES DE RENDERIZADO — definidas a nivel de módulo
# para evitar recrearlas en cada rerun de Streamlit
# ─────────────────────────────────────────────────────────────────────────────
ESTADO_PROY_COLORS = {
    "SIN CONTRATAR":                 (C["cian"],        "#e0f7fa"),
    "CONTRATADO EN EJECUCIÓN":       (C["verde_medio"], "#d1fae5"),
    "CONTRATADO SIN ACTA DE INICIO": (C["azul_medio"],  "#dbeafe"),
    "TERMINADO":                     (C["muted"],       "#f1f5f9"),
    "PARA CIERRE":                   (C["cafe"],        "#fef3c7"),
}
ESTADO_CONT_COLORS = {
    "EN EJECUCIÓN":  (C["verde_medio"], "#d1fae5"),
    "TERMINADO":     (C["muted"],       "#f1f5f9"),
    "LIQUIDADO":     (C["azul_medio"],  "#dbeafe"),
    "SUSPENDIDO":    (C["naranja_osc"], "#ffedd5"),
    "SIN CONTRATO":  (C["cian"],        "#e0f7fa"),
}
CTTO_ESTADO_COLORS = {
    "EN EJECUCIÓN":  (C["verde_medio"],  "#d1fae5"),
    "EJECUTADO":     (C["verde_oscuro"], "#d1fae5"),
    "TERMINADO":     (C["muted"],        "#f1f5f9"),
    "LIQUIDADO":     (C["azul_medio"],   "#dbeafe"),
    "SUSPENDIDO":    (C["naranja_osc"],  "#ffedd5"),
    "RESCINDIDO":    (C["salmon"],       "#fee2e2"),
    "SUSCRITO":      (C["cian"],         "#e0f7fa"),
}

def _pill(texto, color_map, default_fg=None, default_bg=None):
    if not texto:
        return '<span class="proy-pill proy-pill--empty">—</span>'
    eu = texto.strip().upper()
    fg, bg = color_map.get(eu, (default_fg or C["muted"], default_bg or "#f1f5f9"))
    extra = "font-weight:700;" if eu == "SUSPENDIDO" else ""
    return (f'<span class="proy-pill" '
            f'style="background:{bg};color:{fg};border:1px solid {fg}40;{extra}">'
            f'{html.escape(texto)}</span>')

def _fmt_valor(v):
    """Formatea un valor float como moneda COP."""
    if v is None or (isinstance(v, float) and v != v):
        return "—"
    try:
        return f"$ {v:,.0f}"
    except Exception:
        return str(v)

def _valor_a_gradiente(valor, v_min, v_max):
    """Devuelve un color de fondo suave proporcional al valor del contrato."""
    if valor is None or v_max == v_min:
        return "#ffffff"
    ratio = max(0.0, min(1.0, (valor - v_min) / (v_max - v_min)))
    r = int(255 - ratio * (255 - 219))
    g = int(255 - ratio * (255 - 234))
    b = int(255 - ratio * (255 - 254))
    return f"rgb({r},{g},{b})"

def _contratos_panel(bpin_str, df_cttos):
    """Genera el HTML del panel de contratos para un BPIN dado."""
    if df_cttos is None:
        return '<div class="ctto-panel"><div class="ctto-panel-empty">Archivo de contratos no disponible.</div></div>'

    bpin_norm = (
        str(bpin_str).strip()
        .replace(".", "").replace("-", "").replace(",", "").replace(" ", "")
    )
    cttos = df_cttos.filter(pl.col("BPIN") == bpin_norm)

    if cttos.height == 0:
        return '<div class="ctto-panel"><div class="ctto-panel-empty">Sin contratos registrados para este proyecto.</div></div>'

    valores     = [r.get("CONTRATO VALOR TOTAL") for r in cttos.to_dicts()]
    valores_num = [v for v in valores if v is not None and isinstance(v, float) and v == v]
    v_min = min(valores_num) if valores_num else 0
    v_max = max(valores_num) if valores_num else 0

    n      = cttos.height
    header = f"""
    <div class="ctto-panel-header">
        <span class="ctto-panel-title">Contratos</span>
        <span class="ctto-panel-count">{n}</span>
    </div>"""

    rows = ""
    for ctto in cttos.to_dicts():
        valor    = ctto.get("CONTRATO VALOR TOTAL")
        bg_grad  = _valor_a_gradiente(valor, v_min, v_max)
        estado_c = (ctto.get("ESTADO CONTRATO") or "").strip().upper()
        fg_e, bg_e = CTTO_ESTADO_COLORS.get(estado_c, (C["muted"], "#f1f5f9"))
        proceso   = html.escape(ctto.get("NO. PROCESO PRECONTRACTUAL") or "—")
        modalidad = html.escape(ctto.get("MODALIDAD CONTRATACION") or "—")
        tipo      = html.escape(ctto.get("TIPO CONTRATO") or "—")
        objeto    = html.escape(ctto.get("CONTRATO OBJETO") or "—")

        bar_px = 0
        if valor and v_max > v_min:
            bar_px = int(max(6, min(60, (valor - v_min) / (v_max - v_min) * 60)))
        elif valor:
            bar_px = 60

        rows += f"""<tr style="background:{bg_grad}">
            <td class="ctto-col1"><span class="ctto-proceso">{proceso}</span></td>
            <td style="font-size:0.73rem;color:{C['text']}">{modalidad}</td>
            <td style="font-size:0.73rem;color:{C['muted']}">{tipo}</td>
            <td>
                <div class="ctto-valor-wrap">
                    <span class="ctto-valor">{_fmt_valor(valor)}</span>
                    <div class="ctto-valor-bar" style="width:{bar_px}px"></div>
                </div>
            </td>
            <td><span class="ctto-estado-pill" style="background:{bg_e};color:{fg_e};border:1px solid {fg_e}40">{html.escape(ctto.get("ESTADO CONTRATO") or "—")}</span></td>
            <td><div class="ctto-objeto">{objeto}</div></td>
        </tr>"""

    tabla = f"""
    <div style="border-radius:10px;overflow:hidden;box-shadow:0 1px 10px rgba(0,40,90,0.10);">
    <table class="ctto-table">
    <thead><tr>
        <th class="ctto-col1">No. proceso</th>
        <th>Modalidad</th><th>Tipo</th>
        <th>Valor total</th><th>Estado</th>
        <th>Objeto del contrato</th>
    </tr></thead>
    <tbody>{rows}</tbody>
    </table></div>"""
    return f'<div class="ctto-panel">{header}{tabla}</div>'

# ─────────────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────────────
tab_resumen, tab_proyectos, tab_evaluacion, tab_comunicaciones, tab_exportar = st.tabs([
    "Resumen por entidad",
    "Todos los proyectos",
    "Evaluación del modelo",
    "Comunicaciones",
    "Exportar",
])

# HITO key map para badge_html
HITO_KEY_MAP = {
    "clasi_1": "hito_1_val",
    "clasi_2": "hito_2_val",
    "clasi_3": "hito_3_val",
    "clasi_4": "hito_4_val",
    "clasi_5": "hito_5_val",
}

# ── TAB 1: Tabla resumen ──────────────────────────────────────────────────────
with tab_resumen:
    def hito_cell(dias_val, entidad, clasi_key):
        if dias_val is None or (isinstance(dias_val, float) and dias_val != dias_val):
            return "<td class='null-cell'>—</td>"
        clasi = clasi_por_entidad[entidad].get(clasi_key)
        hito_k = HITO_KEY_MAP.get(clasi_key)
        return f"<td><span class='dias-val'>{dias_val:.1f} d</span>{badge_html(clasi, hito_k)}</td>"

    rows_html = ""
    for row in agrupacion.to_dicts():
        e    = row["ENTIDAD O SECRETARIA"]
        susp = int(row["Suspendidos"]) if row["Suspendidos"] else 0
        pc   = int(row["Para cierre"]) if row["Para cierre"] else 0
        rows_html += f"""<tr>
            <td class="entidad-name">{e}</td>
            {hito_cell(row['Hito 1 (días)'], e, 'clasi_1')}
            {hito_cell(row['Hito 2 (días)'], e, 'clasi_2')}
            {hito_cell(row['Hito 3 (días)'], e, 'clasi_3')}
            {hito_cell(row['Hito 4 (días)'], e, 'clasi_4')}
            {hito_cell(row['Hito 5 (días)'], e, 'clasi_5')}
            <td style="text-align:center;font-weight:500">{susp}</td>
            <td style="text-align:center;font-weight:500">{pc}</td>
            <td class="col-total">{int(row['Total'])}</td>
        </tr>"""

    st.markdown(f"""
    <table class="summary-table">
    <thead><tr>
        <th>Entidad / Secretaría</th>
        {th("Sin contratar<br>sin apertura", "Hito 1 · Sin contratar sin apertura",
            "Promedio de días entre la <b>Fecha de aprobación</b> y la <b>Fecha de corte GESPROY</b>.<br><br>Condición: Estado = SIN CONTRATAR y sin fecha de apertura.")}
        {th("Sin contratar<br>con apertura", "Hito 2 · Sin contratar con apertura",
            "Promedio de días entre la <b>Fecha de apertura del primer proceso</b> y la <b>Fecha de acta de inicio</b>.<br><br>Condición: Estado = SIN CONTRATAR con fecha de apertura registrada.")}
        {th("Contratado<br>sin acta de inicio", "Hito 3 · Contratado sin acta de inicio",
            "Promedio de días entre la <b>Fecha de suscripción</b> y la <b>Fecha de corte GESPROY</b>.<br><br>Condición: Estado = CONTRATADO SIN ACTA DE INICIO.")}
        {th("En ejecución<br>rezagado", "Hito 4 · En ejecución rezagado",
            "Meses entre el <b>Horizonte del proyecto</b> y la <b>Fecha de corte GESPROY</b>.<br><br>Condición: Estado = CONTRATADO EN EJECUCIÓN, CPI = 0, SPI = 0 y horizonte vencido.")}
        {th("Proyectos<br>terminados", "Hito 5 · Proyectos terminados",
            "Promedio de días entre la <b>Fecha de finalización</b> y la <b>Fecha de corte GESPROY</b>.<br><br>Condición: Fecha de finalización registrada.")}
        {th("Suspendidos", "Proyectos suspendidos", "Conteo de proyectos cuyo <b>Estado contrato</b> = SUSPENDIDO.")}
        {th("Para cierre", "Proyectos para cierre", "Conteo de proyectos con Estado = PARA CIERRE.")}
        <th class="col-total">Total</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)

    # ── Sección de detalle por hito (debajo de la tabla resumen) ─────────────
    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)
    st.markdown("<div class='section-heading'>Detalle por hito</div>", unsafe_allow_html=True)
    st.markdown(
        "<div style='font-size:0.78rem;color:#6b7280;margin-bottom:0.8rem'>"
        "Selecciona un hito para ver el detalle de todos los proyectos que lo tienen activo, "
        "ordenados de mayor a menor tiempo.</div>",
        unsafe_allow_html=True,
    )

    sel_hito_resumen = st.selectbox(
        "Hito a detallar",
        list(HITOS.keys()),
        key="sel_hito_resumen",
        label_visibility="collapsed",
    )
    sel_hito_col_r, sel_clasi_col_r = HITOS[sel_hito_resumen]
    hito_key_detalle = HITO_KEY_MAP.get(sel_clasi_col_r, None)

    # Incluir fechas para el tooltip de cálculo
    DATE_COLS_DET = [
        "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
        "FECHA SUSCRIPCION", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
        "FECHA DE FINALIZACIÓN", "FECHA DE CORTE GESPROY",
    ]
    df_det = (
        df_f
        .filter(~pl.col(sel_hito_col_r).is_null())
        .select(
            "ENTIDAD O SECRETARIA", "BPIN", "NOMBRE PROYECTO", "ESTADO PROYECTO",
            sel_hito_col_r, sel_clasi_col_r,
            *DATE_COLS_DET,
        )
        .sort(["ENTIDAD O SECRETARIA", sel_hito_col_r], descending=[False, True])
    )

    def _fmt_date(val):
        """Formatea una fecha como DD/MM/YYYY o devuelve '—'."""
        if val is None:
            return "—"
        try:
            return val.strftime("%d/%m/%Y")
        except Exception:
            return str(val)

    HITO_CALC_META = {
        "hito_1_val": (
            "Fecha aprobación",    "FECHA APROBACIÓN PROYECTO",
            "Fecha corte GESPROY", "FECHA DE CORTE GESPROY",
            "Días desde la aprobación del proyecto hasta el corte, sin proceso de contratación abierto.",
        ),
        "hito_2_val": (
            "Fecha apertura proceso", "FECHA DE APERTURA DEL PRIMER PROCESO",
            "Fecha acta de inicio",   "FECHA ACTA INICIO",
            "Días desde la apertura del primer proceso hasta el acta de inicio.",
        ),
        "hito_3_val": (
            "Fecha suscripción",   "FECHA SUSCRIPCION",
            "Fecha corte GESPROY", "FECHA DE CORTE GESPROY",
            "Días desde la suscripción del contrato hasta el corte, sin acta de inicio.",
        ),
        "hito_4_val": (
            "Horizonte del proyecto", "HORIZONTE DEL PROYECTO",
            "Fecha corte GESPROY",    "FECHA DE CORTE GESPROY",
            "Días de retraso sobre el horizonte (CPI=0, SPI=0). El resultado se muestra en meses.",
        ),
        "hito_5_val": (
            "Fecha finalización",  "FECHA DE FINALIZACIÓN",
            "Fecha corte GESPROY", "FECHA DE CORTE GESPROY",
            "Días entre la fecha de finalización registrada y el corte.",
        ),
    }

    def _dias_tooltip(r, hito_col):
        meta = HITO_CALC_META.get(hito_col)
        if not meta:
            return ""
        lbl_a, col_a, lbl_b, col_b, nota = meta
        fecha_a = _fmt_date(r.get(col_a))
        fecha_b = _fmt_date(r.get(col_b))
        dias_v  = r.get(hito_col)
        dias_display = f"{dias_v:.0f} días" if dias_v is not None else "—"
        es_h4 = hito_col == "hito_4_val"
        resultado_label = f"{dias_v/30:.1f} meses ({dias_display})" if es_h4 and dias_v else dias_display
        return (
            f'<div class="dias-tip-box">'
            f'  <div class="dias-tip-title">Cálculo del hito</div>'
            f'  <div class="dias-tip-row"><span class="dias-tip-lbl">{lbl_b}</span>'
            f'    <span class="dias-tip-val">{fecha_b}</span></div>'
            f'  <div class="dias-tip-op">menos (−)</div>'
            f'  <div class="dias-tip-row"><span class="dias-tip-lbl">{lbl_a}</span>'
            f'    <span class="dias-tip-val">{fecha_a}</span></div>'
            f'  <div class="dias-tip-sep"></div>'
            f'  <div class="dias-tip-result">= &nbsp;{resultado_label}</div>'
            f'  <div class="dias-tip-nota">{nota}</div>'
            f'</div>'
        )

    if df_det.height == 0:
        st.info("No hay proyectos con valor en este hito para los filtros seleccionados.")
    else:
        for entidad in df_det["ENTIDAD O SECRETARIA"].unique().sort().to_list():
            sub  = df_det.filter(pl.col("ENTIDAD O SECRETARIA") == entidad)
            prom = sub[sel_hito_col_r].mean()
            n    = sub.height
            prom_str = f"{prom:.1f} días" if prom is not None else "—"

            with st.expander(f"{entidad}   ·   {n} proyecto(s)   ·   Promedio: {prom_str}", expanded=False):
                det_rows = ""
                for r in sub.to_dicts():
                    dias_v   = r[sel_hito_col_r]
                    clasi_v  = r[sel_clasi_col_r]
                    dias_str = f"{dias_v:.1f} d" if dias_v is not None else "—"
                    _row_cls_map = {
                        "badge-green":  "row-green",
                        "badge-yellow": "row-yellow",
                        "badge-orange": "row-orange",
                        "badge-black":  "row-black",
                    }
                    _cls_badge_map = {
                        "0-100": "badge-green",  "0-30":  "badge-green",  "0-1":   "badge-green",
                        "101-150": "badge-yellow","31-45": "badge-yellow", "1.1-3": "badge-yellow",
                        "151-180": "badge-orange","46-60": "badge-orange", "3.1-6": "badge-orange",
                        ">180":    "badge-black", ">60":   "badge-black",  ">6":    "badge-black",
                    }
                    badge_cls = _cls_badge_map.get(str(clasi_v), "badge-yellow") if clasi_v else ""
                    row_cls   = _row_cls_map.get(badge_cls, "")
                    tooltip   = _dias_tooltip(r, sel_hito_col_r)
                    det_rows += f"""<tr class="{row_cls}">
                        <td><span class="bpin-tag">{r['BPIN'] or '—'}</span></td>
                        <td style="font-size:0.81rem">{r['NOMBRE PROYECTO'] or '—'}</td>
                        <td><span class="estado-tag">{r['ESTADO PROYECTO'] or '(Sin estado)'}</span></td>
                        <td>
                          <div class="dias-tip-wrap">
                            <span class="dias-val-link">{dias_str}</span>
                            {tooltip}
                          </div>
                        </td>
                        <td>{badge_html(clasi_v, hito_key_detalle)}</td>
                    </tr>"""
                st.markdown(f"""
                <table class="detail-table">
                <thead><tr>
                    <th>BPIN</th><th>Nombre del proyecto</th><th>Estado</th>
                    <th>Días <span style="font-size:0.58rem;font-weight:500;opacity:0.7">(pasar el cursor)</span></th>
                    <th>Clasificación</th>
                </tr></thead>
                <tbody>{det_rows}</tbody>
                </table>""", unsafe_allow_html=True)

# ── TAB 2: Todos los proyectos ────────────────────────────────────────────────
with tab_proyectos:
    st.markdown("<div class='section-heading'>Todos los proyectos</div>", unsafe_allow_html=True)

    # ── CSS de contratos (inyectado una sola vez en el tab) ───────────────────
    st.markdown(f"""
    <style>
    /* ── Tabla de proyectos ── */
    .proy-table {{
        width: 100%; border-collapse: collapse; font-size: 0.83rem;
        background: #ffffff; border-radius: 12px; overflow: hidden;
        box-shadow: 0 2px 20px rgba(0,40,90,0.10);
    }}
    .proy-table thead tr {{ background: {C['azul_oscuro']}; color: white; }}
    .proy-table th {{
        padding: 0.85rem 1rem; font-family: 'Montserrat', sans-serif;
        font-size: 0.62rem; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.8px; text-align: left; white-space: nowrap;
    }}
    .proy-table td {{
        padding: 0.85rem 1rem; border-bottom: 1px solid {C['border']};
        vertical-align: middle;
    }}
    .proy-table tbody tr:last-child td {{ border-bottom: none; }}
    .proy-table tbody tr.proy-data-row:hover td {{
        background: #eef5ff !important; transition: background 0.15s;
    }}
    .proy-ent    {{ font-weight:700; font-size:0.8rem; color:{C['azul_oscuro']}; white-space:nowrap; }}
    .proy-nombre {{ font-size:0.82rem; color:{C['text']}; line-height:1.5; }}
    .proy-pill {{
        display:inline-block; font-size:0.68rem; padding:4px 11px;
        border-radius:20px; font-weight:600; white-space:nowrap;
        font-family:'Montserrat',sans-serif;
    }}
    .proy-pill--empty {{ color:{C['muted']}; font-weight:400; }}

    /* ── Botón expandir contratos ── */
    .ctto-toggle {{
        display: inline-flex; align-items: center; gap: 6px;
        background: {C['azul_oscuro']}0f;
        border: 1.5px solid {C['azul_oscuro']}28;
        color: {C['azul_oscuro']}; border-radius: 8px;
        padding: 5px 12px; font-size: 0.68rem; font-weight: 700;
        cursor: pointer; white-space: nowrap; user-select: none;
        font-family: 'Montserrat', sans-serif;
        transition: background 0.15s, border-color 0.15s;
    }}
    .ctto-toggle:hover {{
        background: {C['azul_medio']}18;
        border-color: {C['azul_medio']}55;
        color: {C['azul_medio']};
    }}
    .ctto-toggle.open {{
        background: {C['azul_oscuro']};
        border-color: {C['azul_oscuro']};
        color: white;
    }}
    .ctto-arrow {{ font-size:0.55rem; transition: transform 0.2s; line-height:1; }}
    .ctto-toggle.open .ctto-arrow {{ transform: rotate(90deg); }}

    /* ── Fila contenedor de contratos ── */
    .ctto-detail-row {{ display: none; }}
    .ctto-detail-row.visible {{ display: table-row; }}
    .ctto-detail-row td {{
        padding: 0 !important;
        border: none !important;
        border-bottom: 3px solid {C['azul_oscuro']}25 !important;
    }}

    /* ── Panel interior ── */
    .ctto-panel {{
        padding: 1.1rem 1.4rem 1.2rem 1.4rem;
        background: linear-gradient(180deg, #edf3fb 0%, #f4f8fd 100%);
        border-left: 4px solid {C['azul_medio']};
    }}
    .ctto-panel-header {{
        display: flex; align-items: center; gap: 0.7rem;
        margin-bottom: 0.9rem;
    }}
    .ctto-panel-title {{
        font-family: 'Montserrat', sans-serif; font-size: 0.7rem;
        font-weight: 800; text-transform: uppercase; letter-spacing: 1.2px;
        color: {C['azul_oscuro']};
    }}
    .ctto-panel-count {{
        background: {C['azul_medio']}; color: white;
        border-radius: 20px; padding: 1px 9px;
        font-size: 0.62rem; font-weight: 700;
        font-family: 'DM Mono', monospace;
    }}
    .ctto-panel-empty {{
        font-size: 0.8rem; color: {C['muted']}; font-style: italic;
        padding: 0.5rem 0;
    }}

    /* ── Tabla de contratos ── */
    .ctto-table {{
        width: 100%; border-collapse: collapse; font-size: 0.77rem;
    }}
    .ctto-table thead tr {{
        background: {C['azul_medio']};
    }}
    .ctto-table th {{
        padding: 0.55rem 1rem;
        font-family: 'Montserrat', sans-serif; font-size: 0.59rem;
        font-weight: 700; text-transform: uppercase; letter-spacing: 0.8px;
        color: rgba(255,255,255,0.95); text-align: left;
        white-space: nowrap; border: none;
    }}
    .ctto-table td {{
        padding: 0.5rem 1rem;
        vertical-align: middle; border: none;
        border-bottom: 1px solid rgba(0,0,0,0.06);
    }}
    .ctto-table tbody tr:last-child td {{ border-bottom: none; }}
    .ctto-table tbody tr:hover td     {{ filter: brightness(0.97); }}
    /* Primera columna — clase dedicada para máxima especificidad */
    td.ctto-col1, th.ctto-col1 {{
        padding-left: 1.4rem !important;
    }}

    /* ── Celda valor ── */
    .ctto-valor-wrap {{
        display: inline-block;
    }}
    .ctto-valor {{
        font-family: 'DM Mono', monospace; font-weight: 800;
        font-size: 0.82rem; white-space: nowrap;
        color: {C['azul_oscuro']}; display: block;
    }}
    .ctto-valor-bar {{
        display: block; height: 3px; border-radius: 2px; margin-top: 4px;
        background: {C['azul_medio']}; opacity: 0.4;
    }}

    /* ── Pill estado contrato ── */
    .ctto-estado-pill {{
        display: inline-block; font-size: 0.63rem; padding: 3px 9px;
        border-radius: 12px; font-weight: 700; white-space: nowrap;
        font-family: 'Montserrat', sans-serif; letter-spacing: 0.3px;
    }}

    /* ── Objeto (texto largo) ── */
    .ctto-objeto {{
        font-size: 0.73rem; color: {C['text']}; line-height: 1.5;
        max-width: 320px;
    }}
    .ctto-proceso {{
        font-family: 'DM Mono', monospace; font-size: 0.72rem;
        color: {C['azul_medio']}; font-weight: 600; white-space: nowrap;
    }}
    </style>
    """, unsafe_allow_html=True)

    # ── Banner de filtros ─────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:0.6rem;
        background:{C['azul_oscuro']}0d;border:1px solid {C['azul_oscuro']}22;
        border-left:3px solid {C['cian']};border-radius:6px;
        padding:0.5rem 0.85rem;margin-bottom:0.75rem">
        <span style="font-size:0.78rem;color:{C['azul_oscuro']};font-weight:600;
                     font-family:'Montserrat',sans-serif;letter-spacing:0.3px">
            Filtros disponibles:
        </span>
        <span style="font-size:0.75rem;color:{C['muted']}">
            entidad &nbsp;·&nbsp; estado del proyecto &nbsp;·&nbsp;
            estado del contrato &nbsp;·&nbsp; BPIN &nbsp;·&nbsp; nombre del proyecto
        </span>
        <span style="margin-left:auto;font-size:0.7rem;color:{C['cian']};font-weight:600;white-space:nowrap">
            Puedes combinarlos
        </span>
    </div>
    """, unsafe_allow_html=True)

    fc1, fc2, fc3 = st.columns([2, 1.4, 1.4])
    with fc1:
        busqueda = st.text_input("busqueda_proy", placeholder="Buscar por BPIN o nombre…",
                                 label_visibility="collapsed")
    with fc2:
        entidades_proy = sorted(df_f["ENTIDAD O SECRETARIA"].drop_nulls().unique().to_list())
        sel_ent_proy   = st.multiselect("Entidad", entidades_proy,
                                        placeholder="Todas las entidades",
                                        label_visibility="collapsed")
    with fc3:
        estados_proy_opts = sorted(df_f["ESTADO PROYECTO"].drop_nulls().unique().to_list())
        sel_est_proy      = st.multiselect("Estado proyecto", estados_proy_opts,
                                           placeholder="Todos los estados",
                                           label_visibility="collapsed")
    fc4, fc5 = st.columns([1.4, 4.4])
    with fc4:
        estados_cont_opts = sorted(df_f["ESTADO CONTRATO"].drop_nulls().unique().to_list())
        sel_cont_proy     = st.multiselect("Estado contrato", estados_cont_opts,
                                           placeholder="Todos los contratos",
                                           label_visibility="collapsed")

    # ── Filtrar ───────────────────────────────────────────────────────────────
    df_proy = df_f.select(
        "ENTIDAD O SECRETARIA", "BPIN", "NOMBRE PROYECTO",
        "ESTADO PROYECTO", "ESTADO CONTRATO",
    )
    if busqueda:
        term = busqueda.strip().lower()
        df_proy = df_proy.filter(
            pl.col("NOMBRE PROYECTO").str.to_lowercase().str.contains(term, literal=True) |
            pl.col("BPIN").cast(pl.Utf8).str.to_lowercase().str.contains(term, literal=True)
        )
    if sel_ent_proy:
        df_proy = df_proy.filter(pl.col("ENTIDAD O SECRETARIA").is_in(sel_ent_proy))
    if sel_est_proy:
        df_proy = df_proy.filter(pl.col("ESTADO PROYECTO").is_in(sel_est_proy))
    if sel_cont_proy:
        df_proy = df_proy.filter(pl.col("ESTADO CONTRATO").is_in(sel_cont_proy))

    df_proy = df_proy.sort(["ENTIDAD O SECRETARIA", "NOMBRE PROYECTO"])
    n_proy  = df_proy.height

    # ── Estado de contratos ───────────────────────────────────────────────────
    hay_contratos = df_contratos is not None and df_contratos.height > 0

    # ── Diagnóstico de contratos ──────────────────────────────────────────────
    with st.expander("Verificación del archivo de contratos", expanded=not hay_contratos):
        if _cttos_bytes is None:
            st.error("No se pudo descargar el archivo de contratos desde GitHub. Intenta subirlo manualmente desde el panel izquierdo.")
        elif df_contratos is None:
            st.error("El archivo se descargó pero no pudo leerse correctamente. Verifica que sea el reporte de contratos exportado desde GESPROY.")
            st.caption(_cttos_diag)
        else:
            # Resumen en lenguaje sencillo
            bpins_matriz = set(
                str(b).strip().replace(".", "").replace(",", "").replace(" ", "").replace("-", "")
                for b in df_f["BPIN"].drop_nulls().to_list()
            )
            bpins_cttos = set(df_contratos["BPIN"].drop_nulls().to_list())
            en_comun    = bpins_matriz & bpins_cttos
            sin_ctto    = bpins_matriz - bpins_cttos

            col_a, col_b, col_c = st.columns(3)
            col_a.metric("Contratos cargados",   df_contratos.height)
            col_b.metric("Proyectos con contratos", len(en_comun))
            col_c.metric("Proyectos sin contratos", len(sin_ctto))

            if len(en_comun) == 0:
                st.warning(
                    "Ningún proyecto coincide con los contratos del archivo. "
                    "Verifica que el archivo de contratos corresponda a los mismos proyectos de la matriz."
                )
            elif len(sin_ctto) > 0:
                st.info(
                    f"{len(sin_ctto)} proyecto(s) no tienen contratos registrados en el archivo. "
                    "Esto puede ser normal si son proyectos que aún no han iniciado contratación."
                )

    if not hay_contratos:
        st.warning("No se pudieron cargar los contratos. Puedes subirlos manualmente desde el panel izquierdo.", icon=None)

    st.markdown(
        f"<div style='font-size:0.73rem;color:{C['muted']};margin:0.4rem 0 0.6rem'>"
        f"<strong style='color:{C['azul_oscuro']}'>{n_proy}</strong> proyecto(s) encontrado(s)"
        + (f" &nbsp;·&nbsp; <span style='color:{C['verde_medio']};font-weight:600'>"
           f"{df_contratos.height} contratos cargados</span>" if hay_contratos else "")
        + "</div>",
        unsafe_allow_html=True,
    )

    if n_proy == 0:
        st.info("No hay proyectos que coincidan con los filtros aplicados.")
    else:
        # Construir tabla con filas de contratos intercaladas
        rows_html = ""
        for idx, r in enumerate(df_proy.to_dicts()):
            entidad  = html.escape(r.get("ENTIDAD O SECRETARIA") or "—")
            bpin     = html.escape(str(r.get("BPIN") or "—"))
            nombre   = html.escape(r.get("NOMBRE PROYECTO") or "—")
            est_proy = r.get("ESTADO PROYECTO") or ""
            est_cont = r.get("ESTADO CONTRATO") or ""
            row_id   = f"proy-{idx}"

            es_susp  = est_cont.strip().upper() == "SUSPENDIDO"
            bg_susp  = 'style="background:#fff7ed"' if es_susp else ""

            # Contar contratos para este BPIN (badge en el botón)
            # Normalización idéntica a la aplicada en procesar_contratos()
            bpin_norm = (
                str(bpin).strip()
                .replace(".", "")
                .replace("-", "")
                .replace(",", "")
                .replace(" ", "")
            )
            n_cttos   = 0
            if hay_contratos:
                n_cttos = df_contratos.filter(pl.col("BPIN") == bpin_norm).height
            badge = (f'<span style="background:{C["azul_medio"]};color:white;'
                     f'border-radius:10px;padding:1px 6px;font-size:0.58rem;'
                     f'margin-left:4px;font-weight:700">{n_cttos}</span>'
                     if n_cttos > 0 else
                     f'<span style="background:#e5e7eb;color:{C["muted"]};'
                     f'border-radius:10px;padding:1px 6px;font-size:0.58rem;'
                     f'margin-left:4px">0</span>')

            panel_html = _contratos_panel(bpin, df_contratos)

            rows_html += f"""
            <tr class="proy-data-row" {bg_susp}>
                <td class="proy-ent">{entidad}</td>
                <td><span class="bpin-tag">{bpin}</span></td>
                <td class="proy-nombre">{nombre}</td>
                <td>{_pill(est_proy, ESTADO_PROY_COLORS)}</td>
                <td>{_pill(est_cont, ESTADO_CONT_COLORS)}</td>
                <td style="white-space:nowrap">
                    <span class="ctto-toggle" data-target="{row_id}">
                        <span class="ctto-arrow">▶</span> Contratos{badge}
                    </span>
                </td>
            </tr>
            <tr class="ctto-detail-row" id="{row_id}">
                <td colspan="6">{panel_html}</td>
            </tr>"""

        st.markdown(f"""
        <table class="proy-table">
        <thead><tr>
            <th style="width:150px">Entidad / Secretaría</th>
            <th style="width:120px">BPIN</th>
            <th>Nombre del proyecto</th>
            <th style="width:190px">Estado proyecto</th>
            <th style="width:165px">Estado contrato</th>
            <th style="width:110px">Contratos</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
        </table>
        """, unsafe_allow_html=True)

# ── TAB 4: Evaluación del modelo ──────────────────────────────────────────────
with tab_evaluacion:
    st.markdown("<div class='section-heading'>Evaluación del modelo ejecutor</div>", unsafe_allow_html=True)

    modelo_sel = st.radio(
        "Ejecutor",
        ["Departamento de Sucre", "Descentralizadas"],
        horizontal=True,
        label_visibility="collapsed",
    )
    st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)

    if modelo_sel == "Departamento de Sucre":
        df_eval, cols_eval_ok, eval_errores = procesar_eval_sucre(file_bytes)
        col_entidad    = "ENTIDAD O SECRETARIA"
        label_entidad  = "Entidad / Secretaría"
        contexto_error = "Departamento de Sucre"
    else:
        df_eval, cols_eval_ok, eval_errores = procesar_descentralizadas(file_bytes)
        col_entidad    = "EJECUTOR"
        label_entidad  = "Ejecutor"
        contexto_error = "Descentralizadas"

    if eval_errores:
        _render_eval_errors(eval_errores, contexto_error)

    if df_eval is None or df_eval.height == 0:
        st.info(f"No se encontraron datos de evaluación para «{modelo_sel}».")
    else:
        # ── Colores de calificación (escala 0-100) ──
        def eval_color(score, max_score=100.0):
            ratio = score / max_score if max_score > 0 else 0
            if ratio >= 0.8:   return C["verde_medio"],  "Sobresaliente"
            elif ratio >= 0.6: return C["cian"],         "Satisfactorio"
            elif ratio >= 0.4: return C["naranja"],      "Aceptable"
            else:              return C["salmon"],        "Por mejorar"

        max_score = 100.0

        # ── CSS extra para tabla de evaluación ──
        st.markdown(f"""
        <style>
        .eval-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.83rem;
            background: #ffffff;
            border-radius: 8px;
            overflow: hidden;
            margin-bottom: 0.5rem;
        }}
        .eval-table thead tr {{ background: {C['azul_oscuro']}; color: white; }}
        .eval-table th {{
            padding: 0.65rem 1rem;
            font-size: 0.63rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            text-align: left;
        }}
        .eval-table td {{
            padding: 0.6rem 1rem;
            border-bottom: 1px solid {C['border']};
            vertical-align: middle;
            background: #ffffff;
        }}
        .eval-table tbody tr:nth-child(even) td {{ background: #f7fafd; }}
        .eval-table tbody tr:last-child td {{ border-bottom: none; }}
        .eval-table tbody tr:hover td {{ background: #e8f3ff !important; }}
        .eval-score-pill {{
            font-family: 'DM Mono', monospace;
            font-weight: 700;
            font-size: 0.88rem;
            padding: 3px 12px;
            border-radius: 20px;
            display: inline-block;
        }}
        </style>
        """, unsafe_allow_html=True)

        # ── 4 sub-pestañas, una por calificación ──
        tabs_eval = st.tabs([
            "Desempeño en contratación",
            "Información a tiempo",
            "Ejecución del proyecto",
            "Calidad de la información",
        ])

        for i, (col_cal, label_cal) in enumerate(zip(COLS_EVAL, COLS_EVAL_LABELS)):
            with tabs_eval[i]:
                if col_cal not in cols_eval_ok:
                    st.info(f"No hay datos disponibles para «{label_cal}» debido a errores en el archivo.")
                    continue

                filas = []
                for row in df_eval.sort(col_cal, descending=True, nulls_last=True).to_dicts():
                    nombre = row.get(col_entidad) or "Sin nombre"
                    score  = row.get(col_cal)

                    if score is None:
                        filas.append(f"""<tr>
                            <td class="entidad-name">{nombre}</td>
                            <td style="color:{C['muted']}">—</td>
                        </tr>""")
                        continue

                    color_bar, _ = eval_color(score, max_score)
                    # Fondo suave proporcional al color
                    bg_map = {
                        C["verde_medio"]: "#d1fae5", C["cian"]: "#e0f7fa",
                        C["naranja"]:     "#fff7ed", C["salmon"]: "#fee2e2",
                    }
                    bg = bg_map.get(color_bar, "#f1f5f9")

                    filas.append(f"""<tr>
                        <td class="entidad-name">{nombre}</td>
                        <td>
                            <span class="eval-score-pill" style="background:{bg};color:{color_bar}">
                                {score:.2f}
                            </span>
                        </td>
                    </tr>""")

                if not filas:
                    st.info("No hay registros con calificación para este criterio.")
                else:
                    st.markdown(f"""
                    <table class="eval-table">
                    <thead><tr>
                        <th>{label_entidad}</th>
                        <th>Calificación promedio &nbsp;(escala 0 – {max_score:.0f})</th>
                    </tr></thead>
                    <tbody>{"".join(filas)}</tbody>
                    </table>
                    """, unsafe_allow_html=True)


# ── TAB 5: Exportar ───────────────────────────────────────────────────────────
with tab_exportar:
    st.markdown("<div class='section-heading'>Descargar reporte</div>", unsafe_allow_html=True)
    st.markdown(
        "El archivo incluye dos hojas: **Resumen por entidad** con los promedios por hito y nivel de alerta, "
        "y **Detalle proyectos** con cada proyecto, sus fechas de cálculo, días por hito, semáforo y mensaje.",
        unsafe_allow_html=False,
    )
    st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)

    st.download_button(
        label="Descargar reporte Excel",
        data=generar_excel(df_f, agrupacion, clasi_por_entidad),
        file_name=f"regalias_seguimiento_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── TAB 6: Comunicaciones ─────────────────────────────────────────────────────
with tab_comunicaciones:
    st.markdown("<div class='section-heading'>Comunicaciones</div>", unsafe_allow_html=True)

    # ── Banner explicativo en la parte superior ───────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,{C['azul_oscuro']}08,{C['cian']}12);
        border:1px solid {C['cian']}40; border-left:4px solid {C['cian']};
        border-radius:10px; padding:1rem 1.3rem; margin-bottom:1.2rem;
        font-size:0.8rem; color:{C['text']}; line-height:1.7">
        <div style="font-family:'Montserrat',sans-serif;font-size:0.72rem;font-weight:800;
            text-transform:uppercase;letter-spacing:1px;color:{C['cian']};margin-bottom:0.5rem">
            Cómo funciona
        </div>
        <b>1. Filtra</b> los proyectos por hito, nivel de alerta y entidad. &nbsp;
        <b>2. Selecciona</b> con el checkbox los proyectos que quieres incluir en el correo. &nbsp;
        <b>3. Edita</b> el asunto y el cuerpo del mensaje generado automáticamente. &nbsp;
        <b>4. Copia</b> el asunto y el cuerpo con los botones de copia. &nbsp;
        <b>5. Abre</b> Outlook o Gmail y pégalos en el correo.
    </div>
    """, unsafe_allow_html=True)

    st.markdown(f"""
    <style>
    .com-card {{
        background: {C['white']}; border-radius: 12px;
        padding: 1.3rem 1.5rem; margin-bottom: 1rem;
        box-shadow: 0 1px 8px rgba(0,40,90,0.08);
        border-left: 4px solid {C['azul_medio']};
    }}
    .com-card-title {{
        font-family: 'Montserrat', sans-serif; font-size: 0.7rem;
        font-weight: 800; text-transform: uppercase; letter-spacing: 1.1px;
        color: {C['azul_oscuro']}; margin-bottom: 0.7rem;
    }}
    .com-counter {{
        font-size: 0.72rem; color: {C['muted']}; margin: 0.5rem 0 0.9rem;
    }}
    .com-counter strong {{ color: {C['azul_oscuro']}; }}
    .com-btn-primary {{
        display: inline-flex; align-items: center; gap: 7px;
        background: {C['azul_oscuro']}; color: white !important; border: none;
        border-radius: 8px; padding: 9px 22px; font-size: 0.76rem;
        font-weight: 700; font-family: 'Montserrat', sans-serif;
        cursor: pointer; text-decoration: none !important;
        transition: background 0.15s;
    }}
    .com-btn-primary:hover {{ background: {C['azul_medio']}; }}
    </style>
    """, unsafe_allow_html=True)

    HITO_LABELS_COM = {
        "H1 · Sin contratar sin apertura":    ("hito_1_val", "clasi_1"),
        "H2 · Sin contratar con apertura":    ("hito_2_val", "clasi_2"),
        "H3 · Contratado sin acta de inicio": ("hito_3_val", "clasi_3"),
        "H4 · En ejecución rezagado":         ("hito_4_val", "clasi_4"),
        "H5 · Proyectos terminados":          ("hito_5_val", "clasi_5"),
    }
    CLASI_OPTIONS_COM = {
        "Todos":   None,
        "Verde":   ["0-100", "0-30", "0-1"],
        "Naranja": ["101-150", "31-45", "1.1-3"],
        "Rojo":    ["151-180", "46-60", "3.1-6"],
        "Negro":   [">180", ">60", ">6"],
    }
    HITO_DESCRIPCION_COM = {
        "H1 · Sin contratar sin apertura":    "proyectos sin contratar y sin apertura del proceso precontractual",
        "H2 · Sin contratar con apertura":    "proyectos sin contratar con proceso precontractual abierto",
        "H3 · Contratado sin acta de inicio": "proyectos contratados sin acta de inicio firmada",
        "H4 · En ejecución rezagado":         "proyectos en ejecución con horizonte vencido",
        "H5 · Proyectos terminados":          "proyectos terminados pendientes de cierre",
    }

    # ── PASO 1: Filtros ───────────────────────────────────────────────────────
    st.markdown('<div class="com-card"><div class="com-card-title">Paso 1 &nbsp;·&nbsp; Seleccionar proyectos</div>', unsafe_allow_html=True)
    ca, cb, cc = st.columns([2, 1.4, 1.8])
    with ca:
        com_hito_label = st.selectbox("Hito", list(HITO_LABELS_COM.keys()), key="com_hito", label_visibility="collapsed")
    with cb:
        com_clasi_label = st.selectbox("Clasificación", list(CLASI_OPTIONS_COM.keys()), key="com_clasi", label_visibility="collapsed")
    with cc:
        entidades_com = ["Todas"] + sorted(df_f["ENTIDAD O SECRETARIA"].drop_nulls().unique().to_list())
        com_entidad = st.selectbox("Entidad", entidades_com, key="com_entidad", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    com_hito_col, com_clasi_col = HITO_LABELS_COM[com_hito_label]
    df_com = (
        df_f
        .filter(~pl.col(com_hito_col).is_null())
        .select("ENTIDAD O SECRETARIA", "BPIN", "NOMBRE PROYECTO", "ESTADO PROYECTO", com_hito_col, com_clasi_col)
        .sort(["ENTIDAD O SECRETARIA", com_hito_col], descending=[False, True])
    )
    clasi_vals = CLASI_OPTIONS_COM[com_clasi_label]
    if clasi_vals:
        df_com = df_com.filter(pl.col(com_clasi_col).is_in(clasi_vals))
    if com_entidad != "Todas":
        df_com = df_com.filter(pl.col("ENTIDAD O SECRETARIA") == com_entidad)

    n_com = df_com.height
    st.markdown(
        f'<div class="com-counter"><strong>{n_com}</strong> proyecto(s) encontrado(s)</div>',
        unsafe_allow_html=True,
    )

    if n_com == 0:
        st.info("No hay proyectos con este hito y clasificación. Ajusta los filtros.")
    else:
        # ── PASO 2: Tabla seleccionable ───────────────────────────────────────
        df_com_pd = df_com.to_pandas()
        df_com_pd.insert(0, "Incluir", True)
        df_com_pd = df_com_pd.rename(columns={
            com_hito_col:             "Días",
            com_clasi_col:            "Alerta",
            "ENTIDAD O SECRETARIA":   "Entidad",
            "NOMBRE PROYECTO":        "Nombre del proyecto",
            "ESTADO PROYECTO":        "Estado",
        })

        edited = st.data_editor(
            df_com_pd[["Incluir", "Entidad", "BPIN", "Nombre del proyecto", "Estado", "Días", "Alerta"]],
            column_config={
                "Incluir":             st.column_config.CheckboxColumn("✓", width="small"),
                "Entidad":             st.column_config.TextColumn("Entidad", width="medium"),
                "BPIN":                st.column_config.TextColumn("BPIN", width="small"),
                "Nombre del proyecto": st.column_config.TextColumn("Nombre del proyecto", width="large"),
                "Estado":              st.column_config.TextColumn("Estado", width="medium"),
                "Días":                st.column_config.NumberColumn("Días", format="%.0f", width="small"),
                "Alerta":              st.column_config.TextColumn("Alerta", width="small"),
            },
            hide_index=True,
            use_container_width=True,
            key="com_editor",
        )

        proyectos_sel = edited[edited["Incluir"] == True]
        n_sel = len(proyectos_sel)
        st.markdown(
            f'<div class="com-counter"><strong>{n_sel}</strong> proyecto(s) seleccionado(s) para el correo</div>',
            unsafe_allow_html=True,
        )

        if n_sel > 0:
            # ── PASO 2: Redactar ──────────────────────────────────────────────
            st.markdown(f'<div class="com-card" style="border-left-color:{C["cian"]}"><div class="com-card-title">Paso 2 &nbsp;·&nbsp; Redactar correo</div>', unsafe_allow_html=True)

            col_asunto, col_dest = st.columns([3, 2])
            with col_asunto:
                entidad_display = com_entidad if com_entidad != "Todas" else "varias entidades"
                asunto_default  = f"Seguimiento Regalías · {com_hito_label.split('·')[0].strip()} · {entidad_display}"
                com_asunto = st.text_input("Asunto", value=asunto_default, key="com_asunto")
            with col_dest:
                com_dest = st.text_input("Destinatario(s) — separar con comas", placeholder="correo@entidad.gov.co", key="com_dest")

            # Generar lista de proyectos para el cuerpo
            # Mapa hito → (col_clasi, semaforos_dict, calc_explicacion)
            HITO_CALC_EXPLICACION = {
                "H1 · Sin contratar sin apertura": (
                    "hito_1_val",
                    "Este hito mide los días transcurridos desde la aprobación del proyecto "
                    "hasta la fecha de corte GESPROY, sin que se haya abierto ningún proceso precontractual."
                ),
                "H2 · Sin contratar con apertura": (
                    "hito_2_val",
                    "Este hito mide los días entre la apertura del primer proceso precontractual "
                    "y la firma del acta de inicio del contrato."
                ),
                "H3 · Contratado sin acta de inicio": (
                    "hito_3_val",
                    "Este hito mide los días transcurridos desde la suscripción del contrato "
                    "hasta la fecha de corte GESPROY, sin que se haya firmado el acta de inicio."
                ),
                "H4 · En ejecución rezagado": (
                    "hito_4_val",
                    "Este hito mide los meses de retraso del proyecto respecto a su horizonte "
                    "de ejecución previsto, bajo condición de CPI=0 y SPI=0."
                ),
                "H5 · Proyectos terminados": (
                    "hito_5_val",
                    "Este hito mide los días transcurridos desde la fecha de finalización "
                    "registrada del proyecto hasta la fecha de corte GESPROY."
                ),
            }

            def _lista_proyectos(df_sel):
                """Genera la lista de proyectos con mensaje del semáforo y explicación del cálculo."""
                hito_key_com, calc_exp = HITO_CALC_EXPLICACION[com_hito_label]
                lineas = []
                for _, row in df_sel.iterrows():
                    d       = row["Días"]
                    alerta  = str(row["Alerta"]) if row["Alerta"] == row["Alerta"] else None
                    # Obtener mensaje del semáforo
                    mensaje_sem = ""
                    if alerta and hito_key_com in SEMAFOROS and alerta in SEMAFOROS[hito_key_com]:
                        _, _, mensaje_sem = SEMAFOROS[hito_key_com][alerta]

                    # Formatear días con contexto
                    es_h4 = com_hito_label == "H4 · En ejecución rezagado"
                    if d == d and d is not None:
                        if es_h4:
                            d_str = f"{d/30:.1f} meses ({d:.0f} días)"
                        else:
                            d_str = f"{d:.0f} días"
                    else:
                        d_str = "—"

                    lineas.append(
                        f"  • BPIN {row['BPIN']}  —  {row['Nombre del proyecto']}\n"
                        f"    {mensaje_sem}\n"
                        f"    Tiempo transcurrido: {d_str}.\n"
                        f"    ({calc_exp})"
                    )
                return "\n\n".join(lineas)

            hito_desc  = HITO_DESCRIPCION_COM[com_hito_label]
            lista_txt  = _lista_proyectos(proyectos_sel)
            cuerpo_def = (
                f"Estimados,\n\n"
                f"Por medio del presente, nos permitimos informar que en el marco del seguimiento "
                f"y evaluación de proyectos de regalías, se han identificado {n_sel} proyecto(s) "
                f"con {hito_desc}:\n\n"
                f"{lista_txt}\n\n"
                f"Solicitamos respetuosamente su atención a estos proyectos y la adopción de las "
                f"medidas necesarias para avanzar en su gestión.\n\n"
                f"Quedamos atentos a cualquier inquietud.\n\n"
                f"Cordialmente,\n"
                f"Secretaría Técnica · Regalías\n"
                f"Departamento de Sucre"
            )

            # Key dinámica: cambia con los filtros → Streamlit reinicia el widget
            _body_key = f"com_cuerpo_{com_hito_label}_{com_clasi_label}_{com_entidad}_{n_sel}"

            com_cuerpo = st.text_area(
                "Cuerpo del correo (editable)",
                value=cuerpo_def,
                height=380,
                key=_body_key,
            )

            # ── Botones copiar + abrir correo ─────────────────────────────────
            _dest   = com_dest.strip() if com_dest.strip() else ""
            _mailto = f"mailto:{urllib.parse.quote(_dest)}" if _dest else "mailto:"
            _gmail  = "https://mail.google.com/mail/#compose"
            _verde  = C["verde_oscuro"]

            # json.dumps escapa correctamente todos los caracteres especiales
            # incluyendo \, comillas, saltos de línea, caracteres Unicode, etc.
            # Mucho más seguro que replace("'", "\\'")
            _asunto_js = json.dumps(com_asunto)
            _cuerpo_js = json.dumps(com_cuerpo)
            _dest_js   = json.dumps(com_dest)

            components.html(f"""
            <style>
            .copy-btn {{
                display:inline-flex; align-items:center; gap:6px;
                background:#f1f5f9; border:1.5px solid #cbd5e1;
                color:#1a2332; border-radius:7px; padding:7px 18px;
                font-size:0.73rem; font-weight:700; cursor:pointer;
                font-family:'Montserrat',sans-serif; margin:6px 8px 6px 0;
                transition:all 0.15s;
            }}
            .copy-btn:hover {{ background:#e2e8f0; border-color:#94a3b8; }}
            .copy-btn.copied {{ background:#d1fae5; border-color:#059669; color:#065f46; }}
            .open-btn {{
                display:inline-flex; align-items:center; gap:6px;
                border:none; border-radius:7px; padding:7px 18px;
                font-size:0.73rem; font-weight:700; cursor:pointer;
                font-family:'Montserrat',sans-serif; margin:6px 8px 6px 0;
                text-decoration:none; transition:background 0.15s;
            }}
            .open-outlook {{ background:#003d6c; color:white; }}
            .open-outlook:hover {{ background:#1754ab; color:white; }}
            .open-gmail {{ background:#005931; color:white; }}
            .open-gmail:hover {{ background:#17743d; color:white; }}
            </style>

            <div style="display:flex;flex-wrap:wrap;align-items:center;gap:4px;margin-top:8px">
                <button class="copy-btn" id="btn_dest"   onclick="doCopy('{_dest_js}','btn_dest','Copiar destinatario')">
                    Copiar destinatario
                </button>
                <button class="copy-btn" id="btn_asunto" onclick="doCopy('{_asunto_js}','btn_asunto','Copiar asunto')">
                    Copiar asunto
                </button>
                <button class="copy-btn" id="btn_cuerpo" onclick="doCopy('{_cuerpo_js}','btn_cuerpo','Copiar cuerpo')">
                    Copiar cuerpo
                </button>
                <a href="{_mailto}" class="open-btn open-outlook" target="_blank">Abrir Outlook</a>
                <a href="{_gmail}"  class="open-btn open-gmail"   target="_blank">Abrir Gmail</a>
            </div>

            <script>
            function doCopy(text, btnId, label) {{
                var btn = document.getElementById(btnId);
                function confirm() {{
                    btn.innerText = '✓ Copiado';
                    btn.classList.add('copied');
                    setTimeout(function() {{
                        btn.innerText = label;
                        btn.classList.remove('copied');
                    }}, 2000);
                }}
                if (navigator.clipboard && navigator.clipboard.writeText) {{
                    navigator.clipboard.writeText(text).then(confirm).catch(function() {{
                        fallback(text); confirm();
                    }});
                }} else {{
                    fallback(text); confirm();
                }}
            }}
            function fallback(text) {{
                var ta = document.createElement('textarea');
                ta.value = text; ta.style.position='fixed'; ta.style.opacity='0';
                document.body.appendChild(ta); ta.focus(); ta.select();
                try {{ document.execCommand('copy'); }} catch(e) {{}}
                document.body.removeChild(ta);
            }}
            </script>
            """, height=60, scrolling=False)

            st.markdown("</div>", unsafe_allow_html=True)
